# shipment_operations.py
import json
import pandas as pd
import logging
from pathlib import Path

from PyQt6.QtWidgets import QMessageBox, QFileDialog, QDialog, QHeaderView
from PyQt6.QtCore import Qt
import database
from db_connection import get_db_type, _release_connection
import config
import utils
from models import Shipment, ShipmentItem, GroupShipment
from dialogs import RenameDialog

logger = logging.getLogger(__name__)


class ShipmentOperations:
    def __init__(self, main_window):
        self.main_window = main_window

    def start_new_shipment(self):
        # Проверяем, не выполняется ли уже операция создания поставки
        if hasattr(self, '_creating_shipment_in_progress') and self._creating_shipment_in_progress:
            return  # Операция уже выполняется, выходим

        # Устанавливаем флаг, что операция в процессе
        self._creating_shipment_in_progress = True
        
        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self.main_window, "Выберите файл Excel с составом поставки", "", "Excel (*.xlsx *.xls)"
            )
            if not file_path:
                return

            try:
                # Определяем тип поставки по структуре файла
                df_temp = pd.read_excel(file_path, header=None)
                
                # Find the header row
                header_row = 0
                for i in range(min(10, len(df_temp))):
                    row_values = df_temp.iloc[i].dropna().astype(str).str.lower().tolist()
                    if any('штрихкод' in str(val).lower() or 'артикул' in str(val).lower() for val in row_values):
                        header_row = i
                        break
                
                # Read with identified header
                # skiprows=header_row пропускает строки ПЕРЕД заголовком
                # header=0 означает, что первая строка прочитанных данных используется как заголовок
                header_df = pd.read_excel(file_path, skiprows=header_row, header=0)
                
                # Identify the barcode column from the header row
                barcode_col = None
                sku_col_name = None
                for col in header_df.columns:
                    col_str = str(col).lower().strip()
                    if 'штрихкод' in col_str or 'шк' in col_str:
                        barcode_col = col
                    elif 'артикул' in col_str or 'арт' in col_str:
                        sku_col_name = col
                
                # If we can't identify the barcode column by name, use the first column
                if barcode_col is None and len(header_df.columns) > 0:
                    barcode_col = header_df.columns[0]

                # Проверяем, что колонка штрихкодов найдена
                if barcode_col is None:
                    QMessageBox.critical(
                        self.main_window, "Ошибка",
                        "Не найдена колонка со штрихкодами в файле! Проверьте формат файла."
                    )
                    return

                # Now read the file again with the barcode column specified as string to preserve leading zeros
                # skiprows=header_row пропускает строки ПЕРЕД заголовком, чтобы не потерять данные
                dtype_dict = {barcode_col: str}
                df = pd.read_excel(file_path, skiprows=header_row, header=0, dtype=dtype_dict)
                
                # Check if it's a group shipment by counting columns beyond the basic ones
                required_columns = ["Штрихкод", "Артикул", "Количество"]
                basic_columns_present = sum(1 for col in required_columns if any(c.lower() in str(col).lower() for c in ["штрихкод", "артикул", "количество"]))
                
                # If we have more columns than just the basic ones, it's likely a group shipment
                additional_columns = [col for col in df.columns if not any(c.lower() in str(col).lower() for c in ["штрихкод", "шк", "артикул", "арт", "количество", "кол-во"])]
                additional_columns = [col for col in additional_columns if str(col).strip() and not str(col).startswith('Unnamed')]
                
                # Улучшенная логика определения групповой поставки:
                # Групповая поставка - это когда есть более 3 значимых колонок
                # (не считая технические колонки типа Unnamed)
                
                # Определяем значимые колонки (не Unnamed)
                significant_columns = [col for col in df.columns 
                                     if str(col).strip() and not str(col).startswith('Unnamed')]
                
                # Для одиночной поставки должно быть ровно 3 основные колонки:
                # barcode/sku + sku/artikul + quantity/kol-vo
                # Ищем точные названия
                has_barcode = any(c.lower().strip() in ['баркод', 'штрихкод', 'штрих-код', 'шк', 'barcode'] 
                                for c in df.columns)
                has_sku = any(c.lower().strip() in ['артикул', 'арт', 'sku', 'article'] 
                            for c in df.columns)
                has_quantity = any(c.lower().strip() in ['количество', 'кол-во', 'кол', 'quantity', 'qty', 'колич'] 
                                 for c in df.columns)
                
                # Проверяем точные совпадения для базовых колонок
                exact_basic_count = sum([has_barcode, has_sku, has_quantity])
                
                is_group = False
                # Групповая поставка: если есть все 3 основные колонки + дополнительные (>3 всего)
                if exact_basic_count >= 2 and len(significant_columns) > 3:
                    is_group = True
                # Или если есть явные признаки группы: более 4 колонок
                elif len(significant_columns) > 4:
                    is_group = True
                
                if is_group:
                    # This looks like a group shipment - process as group
                    self._process_as_group_shipment(file_path, df, header_row, barcode_col)
                else:
                    # Find quantity column - ищем колонку с количеством
                    quantity_col = None
                    # Сначала пробуем найти по названию
                    for col in df.columns:
                        col_str = str(col).lower().strip()
                        if 'количество' in col_str or 'кол-во' in col_str or 'qty' in col_str or 'quant' in col_str or 'count' in col_str:
                            quantity_col = col
                            break
                    
                    # Если не нашли по названию, используем третью колонку (C = index 2)
                    if quantity_col is None and len(df.columns) > 2:
                        quantity_col = df.columns[2]
                    
                    # Process as regular shipment
                    self._process_as_regular_shipment(file_path, df, barcode_col, sku_col_name, quantity_col)
            
            except Exception as e:
                QMessageBox.critical(self.main_window, "Ошибка", f"Не удалось загрузить файл:\n{e}")
        finally:
            # Сбрасываем флаг после завершения операции
            self._creating_shipment_in_progress = False
    
    def update_group_shipment_from_google_sheets_data(self, result):
        """Обновление групповой поставки из данных Google Sheets.
        Обновляет состав существующих подпоставок, не удаляя их.
        Создаёт новые подпоставки только для колонок, которых ещё нет."""
        df = result['df']
        barcode_col = result['barcode_col']
        sheet_name = result['sheet_name']
        group_shipment = result['group_shipment']
        
        sku_col = None
        for col in df.columns:
            col_lower = str(col).lower().strip()
            if 'артикул' in col_lower or 'арт' in col_lower or 'sku' in col_lower:
                sku_col = col
                break
        
        skip_cols = 2
        if len(df.columns) > 3:
            third_col = df.columns[2]
            # Проверяем, есть ли хотя бы одно числовое (непустое) значение в третьей колонке
            has_numeric = False
            for val in df[third_col]:
                if pd.isna(val):
                    continue
                val_str = str(val).strip()
                if val_str == '' or val_str.lower() in ['nan', 'none']:
                    continue
                try:
                    float(val_str.replace(',', '.'))
                    has_numeric = True
                    break
                except (ValueError, TypeError):
                    continue
            if not has_numeric:
                skip_cols = 3
        
        quantity_cols = list(df.columns[skip_cols:])
        
        if not quantity_cols:
            QMessageBox.warning(self.main_window, "Предупреждение", "Не найдены колонки поставок")
            return
        
        imported_count = 0
        for qty_col in quantity_cols:
            sub_name = str(qty_col).strip()
            if len(sub_name) > 50:
                sub_name = sub_name[:50] + "..."
            if not sub_name:
                sub_name = f"Поставка {quantity_cols.index(qty_col) + 1}"
            
            sub_destination = f"{group_shipment.group_name}::{sub_name}"
            
            # Собираем товары из Google Sheets для этой колонки
            new_items = {}
            for _, row in df.iterrows():
                barcode = str(row[barcode_col]).strip() if barcode_col in df.columns and pd.notna(row.get(barcode_col)) else None
                if not barcode or barcode.lower() in ['nan', 'none', '']:
                    continue
                barcode = barcode.replace(" ", "").replace("-", "").replace("\t", "")
                if not pd.notna(row.get(qty_col)):
                    continue
                try:
                    qty = int(float(row[qty_col]))
                    if qty <= 0:
                        continue
                except (ValueError, TypeError):
                    continue
                sku = ""
                if sku_col and sku_col in df.columns and pd.notna(row.get(sku_col)):
                    sku = str(row[sku_col]).strip()
                new_items[barcode] = (sku if sku else barcode, qty)
            
            if not new_items:
                continue
            
            existing = group_shipment.sub_shipments.get(sub_destination)
            
            if existing:
                # Подпоставка уже есть — обновляем состав, сохраняя коробки и allocated_qty
                self._update_existing_shipment_items(existing, new_items)
            else:
                # Новая подпоставка
                shipment = Shipment(sub_destination, self.main_window.font_size, self.main_window.label_font_size, self.main_window.current_theme)
                shipment.original_destination_name = sub_destination
                shipment.display_name = sub_name
                for barcode, (sku, qty) in new_items.items():
                    shipment.add_shipment_item(barcode, sku, qty)
                group_shipment.add_sub_shipment(sub_destination, shipment)
                self._update_shipment_items_preserving_boxes(sub_destination, shipment)
            
            imported_count += 1
        
        if imported_count > 0:
            self.main_window.group_shipments[group_shipment.group_name] = group_shipment
            group_shipment.invalidate_caches()
            self.main_window.ui_updater.update_ui()
            self.main_window.show_status(f"Групповая поставка '{sheet_name}' обновлена ({imported_count} колонок)", 3000)
        else:
            QMessageBox.warning(self.main_window, "Предупреждение", "Не удалось импортировать товары")
    
    def _update_existing_shipment_items(self, shipment, new_items):
        """Обновляет состав товара существующей подпоставки, сохраняя коробки и allocated_qty.
        new_items: dict[barcode] = (sku, total_qty)"""
        from database import get_db_type
        from db_connection import execute_query, execute_many
        
        db_type = get_db_type()
        use_sqlite = db_type == "sqlite"
        ph = "?" if use_sqlite else "%s"
        
        try:
            shipment_id_result = execute_query(
                f"SELECT id FROM shipments WHERE destination_name = {ph}",
                (shipment.destination_name,),
                fetchone=True
            )
            if not shipment_id_result:
                return
            shipment_id = shipment_id_result[0]
            
            existing_items = execute_query(
                f"SELECT barcode, total_qty FROM shipment_items WHERE shipment_id = {ph}",
                (shipment_id,),
                fetchall=True
            )
            existing_barcodes = {row[0]: row[1] for row in existing_items} if existing_items else {}
            
            new_barcodes = set(new_items.keys())
            old_barcodes = set(existing_barcodes.keys())
            
            # Обновляем shipment_items в памяти
            logger.info(f"_update_existing_shipment_items: {shipment.destination_name} — new_items={len(new_items)}, existing={len(existing_barcodes)}")
            
            # Удаляем товары, которых больше нет в new_items
            barcodes_to_remove = old_barcodes - new_barcodes
            if barcodes_to_remove:
                logger.info(f"  Удаляем товары: {barcodes_to_remove}")
                for barcode in barcodes_to_remove:
                    if barcode in shipment.shipment_items:
                        del shipment.shipment_items[barcode]
            
            for barcode, (sku, qty) in new_items.items():
                if barcode in shipment.shipment_items:
                    old_qty = shipment.shipment_items[barcode].total_qty
                    shipment.shipment_items[barcode].total_qty = qty
                    if sku:
                        shipment.shipment_items[barcode].sku = sku
                    logger.info(f"  Обновляем {barcode}: {old_qty} → {qty}")
                else:
                    shipment.add_shipment_item(barcode, sku, qty)
                    logger.info(f"  Добавляем {barcode}: qty={qty}")
            
            shipment.invalidate_caches()
            
            # Пересчитываем allocated_qty из коробок
            for item in shipment.shipment_items.values():
                item.allocated_qty = sum(b.items.get(item.barcode, 0) for b in shipment.boxes)
            
            # Обновляем БД: удалить устаревшие, обновить существующие, вставить новые
            deletes = []
            updates = []
            inserts = []
            
            # Удаляем товары, которых нет в new_items
            for barcode in barcodes_to_remove:
                if barcode in existing_barcodes:
                    deletes.append((shipment_id, barcode))
            
            for item in shipment.shipment_items.values():
                if item.barcode in existing_barcodes:
                    if item.total_qty != existing_barcodes[item.barcode]:
                        updates.append((item.total_qty, shipment_id, item.barcode))
                else:
                    sku_val = item.sku.encode('utf-8').decode('utf-8') if isinstance(item.sku, str) else item.sku
                    inserts.append((shipment_id, item.barcode, sku_val, item.total_qty, item.allocated_qty))
            
            if deletes:
                execute_many(
                    f"DELETE FROM shipment_items WHERE shipment_id = {ph} AND barcode = {ph}",
                    deletes,
                    template=f"({ph}, {ph})"
                )
                logger.info(f"  Удалено из БД: {len(deletes)} товаров")
            
            if updates:
                execute_many(
                    f"UPDATE shipment_items SET total_qty = {ph}, version = version + 1, updated_at = CURRENT_TIMESTAMP WHERE shipment_id = {ph} AND barcode = {ph}",
                    updates,
                    template=f"({ph}, {ph}, {ph})"
                )
                logger.info(f"  Обновлено в БД: {len(updates)} товаров")
            
            if inserts:
                execute_many(
                    f"INSERT INTO shipment_items (shipment_id, barcode, sku, total_qty, allocated_qty) VALUES {ph} ON CONFLICT (shipment_id, barcode) DO UPDATE SET sku = EXCLUDED.sku, total_qty = EXCLUDED.total_qty, allocated_qty = EXCLUDED.allocated_qty, version = shipment_items.version + 1, updated_at = CURRENT_TIMESTAMP",
                    inserts,
                    template=f"({ph}, {ph}, {ph}, {ph}, {ph})"
                )
                logger.info(f"  Вставлено в БД: {len(inserts)} товаров")
            
            # Обновляем allocated_qty в БД для товаров, которые уже были
            for item in shipment.shipment_items.values():
                if item.barcode in existing_barcodes:
                    execute_query(
                        f"UPDATE shipment_items SET allocated_qty = {ph}, version = version + 1, updated_at = CURRENT_TIMESTAMP WHERE shipment_id = {ph} AND barcode = {ph}",
                        (item.allocated_qty, shipment_id, item.barcode)
                    )
            
            logger.info(f"Обновлена подпоставка {shipment.destination_name}: "
                       f"удалено {len(deletes)}, обновлено {len(updates)}, добавлено {len(inserts)}, "
                       f"товаров в файле: {len(new_barcodes)}, в БД: {len(existing_barcodes)}")
        
        except Exception as e:
            logger.error(f"Ошибка обновления подпоставки {shipment.destination_name}: {e}", exc_info=True)
    
    def _update_shipment_items_preserving_boxes(self, destination_name, shipment):
        """Обновляет total_qty товаров поставки, сохраняя существующие коробки и box_items"""
        from database import get_db_type
        from db_connection import execute_query, execute_many
        from models import Box
        
        db_type = get_db_type()
        use_sqlite = db_type == "sqlite"
        ph = "?" if use_sqlite else "%s"
        
        try:
            shipment_id_result = execute_query(
                f"SELECT id FROM shipments WHERE destination_name = {ph}",
                (destination_name,),
                fetchone=True
            )
            if not shipment_id_result:
                return
            
            shipment_id = shipment_id_result[0]
            
            existing_items = execute_query(
                f"SELECT barcode, total_qty FROM shipment_items WHERE shipment_id = {ph}",
                (shipment_id,),
                fetchall=True
            )
            existing_barcodes = {row[0]: row[1] for row in existing_items} if existing_items else {}
            
            # Загружаем коробки из БД только если они ещё не загружены
            boxes_from_db = execute_query(
                f"SELECT id, box_id FROM boxes WHERE shipment_id = {ph} ORDER BY id",
                (shipment_id,),
                fetchall=True
            )
            
            if not shipment.boxes and boxes_from_db:
                
                box_items_map = {}
                if boxes_from_db:
                    box_ids = [b[0] for b in boxes_from_db]
                    placeholders = ','.join([ph] * len(box_ids))
                    all_box_items = execute_query(
                        f"SELECT box_id, barcode, qty FROM box_items WHERE box_id IN ({placeholders})",
                        tuple(box_ids),
                        fetchall=True
                    )
                    if all_box_items:
                        for bi_row in all_box_items:
                            bid = bi_row[0]
                            if bid not in box_items_map:
                                box_items_map[bid] = {}
                            box_items_map[bid][bi_row[1]] = bi_row[2]
                
                for box_row in boxes_from_db:
                    db_box_id, box_name = box_row[0], box_row[1]
                    box = Box(box_name)
                    if db_box_id in box_items_map:
                        for bc, qty in box_items_map[db_box_id].items():
                            normalized_bc = str(bc).replace(" ", "").replace("-", "").replace("\t", "")
                            if normalized_bc != bc:
                                execute_query(
                                    f"UPDATE box_items SET barcode = {ph} WHERE box_id = {ph} AND barcode = {ph}",
                                    (normalized_bc, db_box_id, bc)
                                )
                            box.items[normalized_bc] = qty
                    shipment.boxes.append(box)
            
            if shipment.boxes:
                shipment.current_box_index = len(shipment.boxes) - 1
            
            for item in shipment.shipment_items.values():
                total_in_boxes = sum(b.items.get(item.barcode, 0) for b in shipment.boxes)
                item.allocated_qty = total_in_boxes
            
            for item in shipment.shipment_items.values():
                if item.allocated_qty > 0:
                    execute_query(
                        f"UPDATE shipment_items SET allocated_qty = {ph}, version = version + 1, updated_at = CURRENT_TIMESTAMP WHERE shipment_id = {ph} AND barcode = {ph}",
                        (item.allocated_qty, shipment_id, item.barcode)
                    )
            
            box_updates = []
            updates = []
            inserts = []
            for item in shipment.shipment_items.values():
                if item.barcode in existing_barcodes:
                    if item.total_qty != existing_barcodes[item.barcode]:
                        updates.append((item.total_qty, shipment_id, item.barcode))
                else:
                    sku_val = item.sku.encode('utf-8').decode('utf-8') if isinstance(item.sku, str) else item.sku
                    inserts.append((shipment_id, item.barcode, sku_val, item.total_qty, item.allocated_qty))
            
            if updates:
                execute_many(
                    f"UPDATE shipment_items SET total_qty = {ph}, version = version + 1, updated_at = CURRENT_TIMESTAMP WHERE shipment_id = {ph} AND barcode = {ph}",
                    updates,
                    template=f"({ph}, {ph}, {ph})"
                )
            
            if inserts:
                execute_many(
                    f"INSERT INTO shipment_items (shipment_id, barcode, sku, total_qty, allocated_qty) VALUES {ph} ON CONFLICT (shipment_id, barcode) DO UPDATE SET sku = EXCLUDED.sku, total_qty = EXCLUDED.total_qty, allocated_qty = EXCLUDED.allocated_qty, version = shipment_items.version + 1, updated_at = CURRENT_TIMESTAMP",
                    inserts,
                    template=f"({ph}, {ph}, {ph}, {ph}, {ph})"
                )
            
            items_to_remove = set(existing_barcodes.keys()) - set(shipment.shipment_items.keys())
            if items_to_remove:
                execute_many(
                    f"DELETE FROM shipment_items WHERE shipment_id = {ph} AND barcode = {ph}",
                    [(shipment_id, bc) for bc in items_to_remove],
                    template=f"({ph}, {ph})"
                )
                for box in shipment.boxes:
                    for bc in items_to_remove:
                        box.items.pop(bc, None)
                
                if boxes_from_db:
                    box_ids_tuple = tuple(b[0] for b in boxes_from_db)
                    box_placeholders = ','.join([ph] * len(box_ids_tuple))
                    for bc in items_to_remove:
                        execute_query(
                            f"DELETE FROM box_items WHERE box_id IN ({box_placeholders}) AND barcode = {ph}",
                            (*box_ids_tuple, bc)
                        )
        except Exception as e:
            self.main_window.logger.error(f"Ошибка обновления items без удаления коробок: {e}", exc_info=True)
    
    def create_shipment_from_google_sheets_data(self, result):
        """Создание поставки из данных Google Sheets"""
        df = result['df']
        barcode_col = result['barcode_col']
        is_group = result['is_group']
        sheet_name = result['sheet_name']
        
        if is_group:
            self._create_group_shipment_from_df(df, barcode_col, sheet_name, is_explicit_group=True)
        else:
            self._create_regular_shipment_from_df(df, barcode_col, sheet_name)
    
    def _create_regular_shipment_from_df(self, df, barcode_col, destination):
        """Создание обычной поставки из DataFrame"""
        # Определяем колонку артикула
        sku_col = None
        for col in df.columns:
            col_lower = str(col).lower().strip()
            if 'артикул' in col_lower or 'арт' in col_lower or 'sku' in col_lower:
                sku_col = col
                break
        
        # Определяем колонку количества
        quantity_col = None
        for col in df.columns:
            col_lower = str(col).lower().strip()
            if 'количество' in col_lower or 'кол-во' in col_lower or 'qty' in col_lower:
                quantity_col = col
                break
        
        if quantity_col is None and len(df.columns) > 2:
            quantity_col = df.columns[2]
        
        # Обрезаем название если слишком длинное
        if len(destination) > 50:
            destination = destination[:50] + "..."
        if not destination:
            destination = "Новая поставка"
        
        # Проверяем уникальность имени
        base_destination = destination
        counter = 1
        while destination in self.main_window.shipments:
            destination = f"{base_destination} ({counter})"
            counter += 1
        
        shipment = Shipment(destination, self.main_window.font_size, self.main_window.label_font_size, self.main_window.current_theme)
        shipment.original_destination_name = destination
        
        for _, row in df.iterrows():
            barcode = str(row[barcode_col]).strip() if barcode_col in df.columns and pd.notna(row.get(barcode_col)) else None
            if not barcode or barcode.lower() in ['nan', 'none', '']:
                continue
            
            # Нормализуем штрихкод: убираем пробелы, дефисы, табуляции
            barcode = barcode.replace(" ", "").replace("-", "").replace("\t", "")
            
            sku = ""
            if sku_col and sku_col in df.columns and pd.notna(row.get(sku_col)):
                sku = str(row[sku_col]).strip()
            
            qty = 1
            if quantity_col and quantity_col in df.columns and pd.notna(row.get(quantity_col)):
                try:
                    qty = int(float(row[quantity_col]))
                except (ValueError, TypeError):
                    qty = 1
            
            shipment.add_shipment_item(barcode, sku if sku else barcode, qty)
        
        if shipment.shipment_items:
            self.main_window.shipment_manager.save_shipment(shipment)
            self.main_window.shipments[destination] = shipment
            self.main_window.current_shipment = shipment
            self.main_window.current_shipment.current_box_index = -1
            self.main_window.ui_updater.update_ui()
            self.main_window.show_status(f"Поставка '{destination}' создана", 3000)
        else:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self.main_window, "Предупреждение", "Не удалось импортировать товары из листа")
    
    def _create_group_shipment_from_df(self, df, barcode_col, group_name, is_explicit_group=False):
        """Создание групповой поставки из DataFrame
        
        Args:
            df: DataFrame с данными
            barcode_col: колонка штрихкодов
            group_name: название группы
            is_explicit_group: True если пользователь явно указал что это групповая поставка
        """
        from models import GroupShipment
        
        # Определяем колонки с количеством (колонки поставок)
        quantity_cols = []
        for col in df.columns:
            col_lower = str(col).lower().strip()
            if any(kw in col_lower for kw in ['количество', 'кол-во', 'qty', 'quantity']):
                quantity_cols.append(col)
        
        # Если пользователь явно указал групповую поставку, но не найдены колонки с "количество"
        # - берём все колонки после первых 2 (штрихкод, артикул), которые содержат числовые данные
        if is_explicit_group and not quantity_cols:
            # Пропускаем первые 2 колонки (штрихкод, артикул)
            # Все остальные колонки - это города с количествами
            for col in df.columns[2:]:
                col_name = str(col).strip()
                # Пропускаем пустые и служебные колонки
                if not col_name or col_name.startswith('Unnamed'):
                    continue
                # Проверяем, есть ли хотя бы одно числовое (непустое) значение в колонке
                has_numeric = False
                for val in df[col]:
                    if pd.isna(val):
                        continue
                    val_str = str(val).strip()
                    if val_str == '' or val_str.lower() in ['nan', 'none']:
                        continue
                    try:
                        float(val_str.replace(',', '.'))
                        has_numeric = True
                        break
                    except (ValueError, TypeError):
                        continue
                if has_numeric:
                    quantity_cols.append(col)
        
        if not quantity_cols:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self.main_window, "Предупреждение", "Не найдены колонки с количеством")
            return
        
        # Определяем колонку артикула
        sku_col = None
        for col in df.columns:
            col_lower = str(col).lower().strip()
            if 'артикул' in col_lower or 'арт' in col_lower or 'sku' in col_lower:
                sku_col = col
                break
        
        # Обрезаем название группы
        if len(group_name) > 50:
            group_name = group_name[:50] + "..."
        if not group_name:
            group_name = "Групповая поставка"
        
        # Проверяем уникальность имени группы
        base_group_name = group_name
        counter = 1
        while group_name in self.main_window.group_shipments:
            group_name = f"{base_group_name} ({counter})"
            counter += 1
        
        group_shipment = GroupShipment(group_name, self.main_window.font_size, self.main_window.label_font_size, self.main_window.current_theme)
        
        # Создаём подпоставки для каждой колонки количества
        for qty_col in quantity_cols:
            # Название подпоставки = название колонки количества
            sub_name = str(qty_col).strip()
            if len(sub_name) > 50:
                sub_name = sub_name[:50] + "..."
            if not sub_name:
                sub_name = f"Поставка {quantity_cols.index(qty_col) + 1}"
            
            sub_destination = f"{group_name}::{sub_name}"
            
            shipment = Shipment(sub_destination, self.main_window.font_size, self.main_window.label_font_size, self.main_window.current_theme)
            shipment.original_destination_name = sub_destination
            shipment.display_name = sub_name
            
            for _, row in df.iterrows():
                barcode = str(row[barcode_col]).strip() if barcode_col in df.columns and pd.notna(row.get(barcode_col)) else None
                if not barcode or barcode.lower() in ['nan', 'none', '']:
                    continue
                
                # Нормализуем штрихкод: убираем пробелы, дефисы, табуляции
                barcode = barcode.replace(" ", "").replace("-", "").replace("\t", "")
                
                # Проверяем количество - если пустое, пропускаем этот товар для данной подпоставки
                if not pd.notna(row.get(qty_col)):
                    continue
                
                try:
                    qty = int(float(row[qty_col]))
                    if qty <= 0:
                        continue
                except (ValueError, TypeError):
                    continue
                
                sku = ""
                if sku_col and sku_col in df.columns and pd.notna(row.get(sku_col)):
                    sku = str(row[sku_col]).strip()
                
                shipment.add_shipment_item(barcode, sku if sku else barcode, qty)
            
            if shipment.shipment_items:
                group_shipment.add_sub_shipment(sub_destination, shipment)
                # Сохраняем каждую подпоставку отдельно
                self.main_window.shipment_manager.save_shipment(shipment)
        
        if group_shipment.sub_shipments:
            self.main_window.group_shipments[group_name] = group_shipment
            self.main_window.current_shipment = None
            self.main_window.ui_updater.update_ui()
            self.main_window.show_status(f"Групповая поставка '{group_name}' создана", 3000)
        else:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self.main_window, "Предупреждение", "Не удалось импортировать товары из листа")
    
    def _process_as_regular_shipment(self, file_path, df, barcode_col=None, sku_col=None, quantity_col=None):
        """Обработка файла как обычной поставки"""
        file_name = Path(file_path).stem
        destination = file_name.replace('_', ' ').replace('-', ' ').strip()

        if len(destination) > 50:
            destination = destination[:50] + "..."

        if not destination:
            destination = "Новая поставка"

        base_destination = destination
        counter = 1
        while destination in self.main_window.shipments:
            destination = f"{base_destination} ({counter})"
            counter += 1

        shipment = Shipment(destination, self.main_window.font_size, self.main_window.label_font_size, self.main_window.current_theme)

        # Устанавливаем оригинальное имя поставки для использования в сессиях и других местах
        shipment.original_destination_name = destination

        for _, row in df.iterrows():
            # Используем определенную колонку для штрихкода или первую колонку
            if barcode_col and barcode_col in df.columns:
                barcode = str(row[barcode_col]).strip()
            else:
                barcode = str(row.iloc[0]).strip() if len(row) > 0 else None
            
            if not barcode or barcode.lower() in ['nan', 'none', '']:
                continue

            # Нормализуем штрихкод: убираем пробелы, дефисы, табуляции
            barcode = barcode.replace(" ", "").replace("-", "").replace("\t", "")

            # Используем определенную колонку для артикула или ставим пустую строку
            sku = ""
            if sku_col and sku_col in df.columns:
                sku_val = row[sku_col]
                if pd.notna(sku_val):
                    sku = str(sku_val).strip()
            
            # Используем определенную колонку для количества или ставим 1
            qty = 1
            if quantity_col and quantity_col in df.columns:
                qty_val = row[quantity_col]
                if pd.notna(qty_val):
                    try:
                        qty = int(float(qty_val))
                    except (ValueError, TypeError):
                        qty = 1

            if barcode:
                shipment.add_shipment_item(barcode, sku if sku else barcode, qty)

        if shipment.shipment_items:
            self.main_window.shipment_manager.save_shipment(shipment)
            self.main_window.shipments[destination] = shipment
            self.main_window.current_shipment = shipment
            self.main_window.update_ui()
            self.main_window.ui_updater.update_shipments_tree()
            QMessageBox.information(self.main_window, "Успех", f"Поставка '{destination}' успешно загружена из файла!")
        else:
            QMessageBox.warning(self.main_window, "Предупреждение", "В файле не найдено товаров для загрузки.")
    
    def _process_as_group_shipment(self, file_path, df, header_row, barcode_col):
        """Обработка файла как групповой поставки"""
        from openpyxl import load_workbook
        
        file_name = Path(file_path).stem
        group_name = file_name.replace('_', ' ').replace('-', ' ').strip()

        if len(group_name) > 50:
            group_name = group_name[:50] + "..."

        if not group_name:
            group_name = "Новая группа поставок"

        base_group_name = group_name
        counter = 1
        while group_name in self.main_window.group_shipments:
            group_name = f"{base_group_name} ({counter})"
            counter += 1

        # Открываем файл через openpyxl для более точного определения заголовков
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # Находим строку заголовка - ищем строку со словом "Штрихкод" или "Артикул"
        # openpyxl использует 1-базовую индексацию (строки начинаются с 1)
        actual_header_row = 1  # По умолчанию первая строка
        for row_idx in range(1, min(11, ws.max_row + 1)):  # Проверяем первые 10 строк, начиная с 1
            row_values = [str(cell.value).lower().strip() if cell.value else '' for cell in ws[row_idx]]
            if any('штрихкод' in v or 'шк' in v or 'артикул' in v or 'арт' in v for v in row_values):
                actual_header_row = row_idx
                break

        # Читаем файл заново с правильной строкой заголовка
        # actual_header_row - это 1-based индекс из openpyxl
        # skiprows=actual_header_row-1 пропускает строки ПЕРЕД заголовком
        # header=0 означает, что первая строка прочитанных данных используется как заголовок
        dtype_dict = {0: str}  # Первая колонка (штрихкоды) как строка
        df = pd.read_excel(file_path, skiprows=actual_header_row - 1, header=0, dtype=dtype_dict)
        
        # Определяем колонки по позициям:
        # A (index 0) - штрихкоды
        # B (index 1) - артикулы
        # C+ (index 2+) - направления (подпоставки)
        barcode_col_local = df.columns[0] if len(df.columns) > 0 else None
        sku_col = df.columns[1] if len(df.columns) > 1 else None

        # Проверяем, что колонки штрихкодов и артикулов найдены
        if barcode_col_local is None:
            QMessageBox.critical(
                self.main_window, "Ошибка",
                "Не найдена колонка со штрихкодами в файле! Проверьте формат файла."
            )
            return

        # Все колонки начиная с третьей (C = index 2) - это направления
        # Игнорируем первые две колонки (A=штрихкод, B=артикул) независимо от их названий
        destination_cols = []
        for i in range(2, len(df.columns)):
            col = df.columns[i]
            col_name = str(col).strip()
            # Пропускаем пустые и служебные колонки
            if not col_name or col_name.startswith('Unnamed'):
                continue
            destination_cols.append(col)

        if not destination_cols:
            QMessageBox.critical(
                self.main_window, "Ошибка",
                "Не найдены колонки с направлениями в файле! Начиная с ячейки C1 должны быть названия направлений."
            )
            return

        group_shipment = GroupShipment(
            group_name,
            self.main_window.font_size,
            self.main_window.label_font_size,
            self.main_window.current_theme
        )

        for destination in destination_cols:
            destination_name = str(destination).strip()
            if not destination_name:
                continue

            # Создаем уникальное имя поставки, добавляя префикс группы
            unique_shipment_name = f"{group_name}::{destination_name}"

            shipment = Shipment(
                unique_shipment_name,
                self.main_window.font_size,
                self.main_window.label_font_size,
                self.main_window.current_theme
            )

            # Устанавливаем красивое отображаемое имя без префикса группы
            shipment.display_name = destination_name
            # Сохраняем оригинальное имя для внутреннего использования
            shipment.original_destination_name = unique_shipment_name

            # Обрабатываем строки с данными (пропускаем заголовок, т.к. pandas уже прочитал его правильно)
            for idx, row in df.iterrows():
                # The barcode is already read as string with proper leading zeros
                barcode_raw = row[barcode_col_local] if pd.notna(row[barcode_col_local]) else ""
                if barcode_raw == "" or pd.isna(barcode_raw):
                    continue

                # Ensure barcode is treated as string to preserve leading zeros
                barcode = str(barcode_raw).strip()

                sku = str(row[sku_col]).strip() if pd.notna(row[sku_col]) else ""

                qty_value = row[destination]
                if pd.isna(qty_value) or str(qty_value).strip() == "":
                    qty = 0
                else:
                    try:
                        qty = int(float(str(qty_value)))
                    except (ValueError, TypeError):
                        qty = 0

                if barcode and qty > 0:
                    if barcode in shipment.shipment_items:
                        shipment.shipment_items[barcode].total_qty += qty
                    else:
                        shipment.shipment_items[barcode] = ShipmentItem(barcode, sku, qty)

            if shipment.shipment_items:
                group_shipment.add_sub_shipment(destination_name, shipment)
                self.main_window.shipment_manager.save_shipment(shipment)

        if not group_shipment.sub_shipments:
            QMessageBox.warning(
                self.main_window, "Предупреждение",
                "В файле не найдено данных для создания поставок. Проверьте формат файла."
            )
            return

        self.main_window.group_shipments[group_name] = group_shipment

        # Обновляем кэш shipment_manager для новой группы
        for shipment in group_shipment.sub_shipments.values():
            self.main_window.shipment_manager.update_cache(shipment)

        # Сбрасываем флаг updating_ui, чтобы обновление не блокировалось
        self.main_window.updating_ui = False

        # Принудительно обновляем дерево поставок для немедленного отображения группы
        self.main_window.ui_updater.update_shipments_tree()
        QMessageBox.information(self.main_window, "Успех", f"Групповая поставка '{group_name}' успешно загружена!")
        self.main_window.ui_updater.update_shipments_tree()
        
        sub_shipment_count = len(group_shipment.sub_shipments)
        total_items = sum(sum(item.total_qty for item in shipment.shipment_items.values())
                         for shipment in group_shipment.sub_shipments.values())
        
        # Статус скрыт, сообщения не отображаются
        # self.main_window.statusBar().showMessage(
        #    f"Создана групповая поставка: {group_name} с {sub_shipment_count} направлениями ({total_items} товаров)",
        #    5000
        # )

    def update_shipment_composition(self):
        if not self.main_window.current_shipment:
            QMessageBox.warning(self.main_window, "Ошибка", "Сначала выберите поставку!")
            return
        
        file_path, _ = QFileDialog.getOpenFileName(
            self.main_window, "Выберите файл Excel с обновленным составом поставки", "", "Excel (*.xlsx *.xls)"
        )
        if not file_path:
            return
        
        try:
            # Сохраняем текущее состояние для возможного отката
            original_boxes = self.main_window.current_shipment.boxes[:]
            original_box_index = self.main_window.current_shipment.current_box_index
            original_items = self.main_window.current_shipment.shipment_items.copy()
            original_removed_items = self.main_window.current_shipment.removed_items.copy()

            df = pd.read_excel(file_path, dtype={"Штрихкод": str})
            required_columns = ["Штрихкод", "Артикул", "Количество"]
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                QMessageBox.critical(
                    self.main_window, "Ошибка",
                    f"В файле отсутствуют обязательные колонки: {', '.join(missing_columns)}"
                )
                return

            # Сохраняем текущие коробки и их содержимое перед обновлением
            old_boxes = self.main_window.current_shipment.boxes[:]
            old_current_box_index = self.main_window.current_shipment.current_box_index

            old_allocated = {}
            old_items = self.main_window.current_shipment.shipment_items.copy()

            # Пересчитываем old_allocated из коробок для точности
            for box in old_boxes:
                for barcode, item_qty in box.items.items():
                    if barcode not in old_allocated:
                        old_allocated[barcode] = 0
                    old_allocated[barcode] += item_qty
            
            new_items = {}
            for _, row in df.iterrows():
                barcode = str(row["Штрихкод"]).strip()
                sku = str(row["Артикул"]).strip()
                
                # Обрабатываем количество с проверкой на NaN
                qty_value = row["Количество"]
                if pd.isna(qty_value) or str(qty_value).strip() == "":
                    qty = 0
                else:
                    try:
                        qty = int(float(str(qty_value)))
                    except (ValueError, TypeError):
                        qty = 0
                
                if barcode in new_items:
                    new_items[barcode].total_qty += qty
                else:
                    new_items[barcode] = ShipmentItem(barcode, sku, qty)
            
            # Добавляем словарь для отслеживания типа изменений
            partial_decrease_items = set()  # Товары, количество которых уменьшилось, но не до нуля
            
            # Сохраняем предыдущие списки для определения типа изменений
            old_removed_items = self.main_window.current_shipment.removed_items.copy()
            self.main_window.current_shipment.removed_items.clear()
            
            updated_items = {}
            
            # Сначала обрабатываем товары, которые остались в поставке
            for barcode, new_item in new_items.items():
                if barcode in old_items:
                    existing_item = old_items[barcode]
                    old_total_qty = existing_item.total_qty
                    old_allocated_qty = old_allocated.get(barcode, 0)

                    # Обновляем информацию о товаре
                    existing_item.total_qty = new_item.total_qty
                    existing_item.sku = new_item.sku

                    # Если новое общее количество меньше, чем было подобрано ранее
                    if new_item.total_qty < old_allocated_qty:
                        # Рассчитываем, сколько штук нужно "убрать" из коробок
                        excess_qty = old_allocated_qty - new_item.total_qty

                        # Добавляем информацию о том, что этот товар требует уменьшения
                        # Товар остается в коробках, но будет выделен оранжевым как требующий уменьшения
                        if barcode in self.main_window.current_shipment.removed_items:
                            self.main_window.current_shipment.removed_items[barcode]['allocated_qty'] += excess_qty
                        else:
                            self.main_window.current_shipment.removed_items[barcode] = {
                                'sku': existing_item.sku,
                                'allocated_qty': excess_qty
                            }
                        # Также добавляем в частичное уменьшение, так как товар остается в поставке, но в меньшем количестве
                        partial_decrease_items.add(barcode)

                    # Проверяем, является ли это частичным уменьшением (количество уменьшилось, но не до нуля)
                    # allocated_qty будет пересчитан из коробок позже
                    if new_item.total_qty < old_total_qty and new_item.total_qty > 0:
                        partial_decrease_items.add(barcode)

                    updated_items[barcode] = existing_item
                else:
                    # Новый товар в поставке
                    updated_items[barcode] = new_item
            
            # Обрабатываем товары, которые были в старой поставке, но исчезли в новой
            for barcode, old_item in old_items.items():
                old_allocated_qty = old_allocated.get(barcode, 0)
                if barcode not in new_items and old_allocated_qty > 0:
                    # Если товар был подобран в коробки, но исчез из поставки, добавляем его в removed_items
                    self.main_window.current_shipment.removed_items[barcode] = {
                        'sku': old_item.sku,
                        'allocated_qty': old_allocated_qty
                    }
                    
                    # Товар полностью исчез из поставки, но остается в коробках для выделения красным
                    # Не удаляем товар из коробок, а просто добавляем в removed_items
            
            self.main_window.current_shipment.shipment_items = updated_items
            
            # Сохраняем информацию о частичном уменьшении в поставке
            self.main_window.current_shipment.partial_decrease_items = partial_decrease_items
            
            # Восстанавливаем коробки и их содержимое после обновления
            # Сохраняем все коробки и их содержимое, но отмечаем проблемные товары
            
            # Сначала пересчитаем allocated_qty на основе содержимого коробок
            # Это нужно для корректного сравнения
            for item in self.main_window.current_shipment.shipment_items.values():
                item.allocated_qty = 0
            
            for box in old_boxes:
                for barcode, item_qty in box.items.items():
                    if barcode in self.main_window.current_shipment.shipment_items:
                        self.main_window.current_shipment.shipment_items[barcode].allocated_qty += item_qty
            
            # Теперь обрабатываем проблемные товары
            for box in old_boxes:
                for barcode, item_qty in box.items.items():
                    if barcode not in self.main_window.current_shipment.shipment_items:
                        # Если товар больше не в поставке, но есть в коробке, добавляем его в removed_items
                        # Это позволяет сохранить собранную коробку, но отметить проблемные товары
                        if barcode not in self.main_window.current_shipment.removed_items:
                            # Получаем SKU из старых данных, если возможно
                            old_sku = old_items.get(barcode, {}).sku if barcode in old_items else "Неизвестно"
                            self.main_window.current_shipment.removed_items[barcode] = {
                                'sku': old_sku,
                                'allocated_qty': item_qty
                            }
                        else:
                            # Увеличиваем количество в removed_items (товар уже был отмечен как проблемный)
                            self.main_window.current_shipment.removed_items[barcode]['allocated_qty'] += item_qty
            
            # Применяем обновленные коробки
            self.main_window.current_shipment.boxes = old_boxes
            self.main_window.current_shipment.current_box_index = old_current_box_index
            
            # Сбрасываем кэши в модели поставки при обновлении состава
            self.main_window.current_shipment.invalidate_caches()

            # Сохраняем поставку, сохраняя содержимое коробок
            self.main_window.shipment_manager.save_shipment(self.main_window.current_shipment, preserve_box_items=True)

            # Воспроизводим звук ПЕРЕД обновлением UI для мгновенного отклика
            utils.play_sound(self.main_window.ok_sound, self.main_window.tone_sound)

            # Обновляем только таблицу поставки и коробки (без перестройки дерева для скорости)
            self.main_window.ui_updater.update_current_components()

        except Exception as e:
            # Восстанавливаем исходное состояние при ошибке
            self.main_window.current_shipment.boxes = original_boxes
            self.main_window.current_shipment.current_box_index = original_box_index
            self.main_window.current_shipment.shipment_items = original_items
            self.main_window.current_shipment.removed_items = original_removed_items

            QMessageBox.critical(self.main_window, "Ошибка", f"Не удалось загрузить файл:\n{e}")

    def delete_shipment(self, shipment_name):
        reply = QMessageBox.question(
            self.main_window, "Подтверждение",
            f"Вы уверены, что хотите удалить поставку «{shipment_name}»?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        try:
            # Принудительно сохраняем все отложенные изменения перед удалением
            self.main_window.force_save_session()

            # Используем функцию из модуля database для корректного удаления поставки и всех связанных данных
            # Сначала проверяем, архивирована ли поставка
            db_type = get_db_type()
            placeholder = "?" if db_type == "sqlite" else "%s"
            
            conn = database.get_connection()
            cursor = conn.cursor()

            cursor.execute(
                f"SELECT archived FROM shipments WHERE destination_name = {placeholder}",
                (shipment_name,)
            )
            result = cursor.fetchone()
            _release_connection(conn)

            if result and result[0]:  # Если поставка архивирована
                success = database.delete_archived_shipment(shipment_name)
            else:  # Если поставка не архивирована
                # Для неархивированных поставок также используем функцию удаления архивированной поставки,
                # так как она корректно удаляет все связанные данные
                success = database.delete_archived_shipment(shipment_name)

            if success:
                if shipment_name in self.main_window.shipments:
                    del self.main_window.shipments[shipment_name]
                
                # Проверяем, есть ли поставка в групповых поставках
                for group_name, group_shipment in self.main_window.group_shipments.items():
                    # Обеспечиваем обратную совместимость: сначала проверяем наличие поставки по ключу (старая схема)
                    if shipment_name in group_shipment.sub_shipments:
                        # Для обратной совместимости с существующими данными
                        del group_shipment.sub_shipments[shipment_name]
                        if not group_shipment.sub_shipments:
                            del self.main_window.group_shipments[group_name]
                        break
                    else:
                        # Для новых данных ищем по original_destination_name
                        shipment_key = None
                        for key, sub_shipment in group_shipment.sub_shipments.items():
                            if (hasattr(sub_shipment, 'original_destination_name') and
                                sub_shipment.original_destination_name == shipment_name):
                                shipment_key = key
                                break
                        
                        if shipment_key:
                            del group_shipment.sub_shipments[shipment_key]
                            if not group_shipment.sub_shipments:
                                del self.main_window.group_shipments[group_name]
                            break
                
                if (self.main_window.current_shipment and
                    self.main_window.current_shipment.destination_name == shipment_name):
                    self.main_window.current_shipment = None
                    
                # Обновляем интерфейс для отображения изменений
                self.main_window.update_ui()
                
                # Принудительно обновляем дерево поставок для гарантии визуального обновления
                if hasattr(self.main_window, 'shipments_tree_widget'):
                    self.main_window.shipments_tree_widget.update()
                    self.main_window.shipments_tree_widget.viewport().update()
                    self.main_window.shipments_tree_widget.updateGeometry()
                
                # Дополнительно принудительно обновляем дерево поставок для гарантии полного обновления
                if hasattr(self.main_window, 'shipments_tree_widget'):
                    self.main_window.shipments_tree_widget.update()
                    self.main_window.shipments_tree_widget.updateGeometry()
                    # Принудительно обновляем размеры колонок дерева
                    header = self.main_window.shipments_tree_widget.header()
                    if header:
                        for col in range(header.count()):
                            header.setSectionResizeMode(col, QHeaderView.ResizeMode.ResizeToContents)
                
                # Статус скрыт, сообщения не отображаются
                # self.main_window.statusBar().showMessage(f"Поставка «{shipment_name}» удалена", 3000)
                
                # Дополнительно обновляем дерево поставок для гарантии мгновенного визуального обновления
                self.main_window.ui_updater.update_shipments_tree()
            else:
                QMessageBox.critical(self.main_window, "Ошибка", "Не удалось удалить поставку из базы данных")
                
        except Exception as e:
            QMessageBox.critical(self.main_window, "Ошибка", f"Не удалось удалить поставку:\n{e}")
            logger.error(f"Ошибка удаления поставки: {e}")

    def rename_shipment(self, old_name):
        existing_names = list(self.main_window.shipments.keys())
        # Добавляем имена из групповых поставок тоже, чтобы избежать конфликта имен
        for group_shipment in self.main_window.group_shipments.values():
            existing_names.extend(list(group_shipment.sub_shipments.keys()))
        
        dialog = RenameDialog(old_name, is_shipment=True, existing_names=existing_names, parent=self.main_window)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        new_name = dialog.get_new_name()
        if not new_name or new_name == old_name:
            return
        try:
            database.execute_query(
                "UPDATE shipments SET destination_name = %s WHERE destination_name = %s",
                (new_name, old_name)
            )
            
            # Проверяем, находится ли поставка в обычных поставках или в групповой поставке
            if old_name in self.main_window.shipments:
                # Обычная поставка
                shipment = self.main_window.shipments.pop(old_name)
                shipment.destination_name = new_name
                # Также обновляем original_destination_name, если оно совпадало со старым именем
                if (hasattr(shipment, 'original_destination_name') and
                    shipment.original_destination_name == old_name):
                    shipment.original_destination_name = new_name
                self.main_window.shipments[new_name] = shipment
            else:
                # Возможно, это подпоставка внутри групповой поставки
                shipment_found = False
                for group_name, group_shipment in self.main_window.group_shipments.items():
                    if old_name in group_shipment.sub_shipments:
                        shipment = group_shipment.sub_shipments.pop(old_name)
                        shipment.destination_name = new_name
                        # Обновляем original_destination_name, если оно совпадало со старым именем
                        if (hasattr(shipment, 'original_destination_name') and
                            shipment.original_destination_name == old_name):
                            shipment.original_destination_name = new_name
                        # Используем новое имя как ключ для подпоставки
                        group_shipment.sub_shipments[new_name] = shipment
                        shipment_found = True
                        break
                    else:
                        # Проверяем по original_destination_name
                        for sub_shipment_key, sub_shipment in group_shipment.sub_shipments.items():
                            if (hasattr(sub_shipment, 'original_destination_name') and
                                sub_shipment.original_destination_name == old_name):
                                shipment = group_shipment.sub_shipments.pop(sub_shipment_key)
                                shipment.destination_name = new_name
                                shipment.original_destination_name = new_name
                                # Используем новое имя как ключ для подпоставки
                                group_shipment.sub_shipments[new_name] = shipment
                                shipment_found = True
                                break
                        if shipment_found:
                            break
            
            # Обновляем имя текущей поставки, если она соответствует старому имени
            if self.main_window.current_shipment and self.main_window.current_shipment.destination_name == old_name:
                self.main_window.current_shipment.destination_name = new_name
                # Также обновляем original_destination_name у текущей поставки, если оно совпадало со старым именем
                if (hasattr(self.main_window.current_shipment, 'original_destination_name') and
                    self.main_window.current_shipment.original_destination_name == old_name):
                    self.main_window.current_shipment.original_destination_name = new_name
            
            self.main_window.update_ui()
            # Статус скрыт, сообщения не отображаются
            # self.main_window.statusBar().showMessage(f"Поставка переименована в «{new_name}»", 3000)
        except Exception as e:
            QMessageBox.critical(self.main_window, "Ошибка", f"Не удалось переименовать поставку:\n{e}")

    def update_group_shipment_composition(self, group_shipment):
        file_path, _ = QFileDialog.getOpenFileName(
            self.main_window, "Выберите файл Excel с обновленным составом групповой поставки", "", "Excel (*.xlsx *.xls)"
        )
        if not file_path:
            return

        # Store original state for rollback if needed
        original_sub_shipments = group_shipment.sub_shipments.copy()

        try:
            # Используем openpyxl для надежного определения заголовка
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            # Находим строку заголовка - ищем строку со словом "Штрихкод" или "Артикул"
            # openpyxl использует 1-базовую индексацию (строки начинаются с 1)
            header_row_1based = 1  # По умолчанию первая строка (1-based)
            for row_idx in range(1, min(11, ws.max_row + 1)):  # Проверяем первые 10 строк, начиная с 1
                row_values = [str(cell.value).lower().strip() if cell.value else '' for cell in ws[row_idx]]
                if any('штрихкод' in v or 'шк' in v or 'артикул' in v or 'арт' in v for v in row_values):
                    header_row_1based = row_idx  # Сохраняем как 1-based индекс
                    break

            # Now read the file with the barcode column specified as string to preserve leading zeros
            # skiprows=header_row_1based-1 пропускает строки ПЕРЕД заголовком
            # header=0 означает, что первая строка прочитанных данных используется как заголовок
            dtype_dict = {0: str}
            df = pd.read_excel(file_path, skiprows=header_row_1based - 1, header=0, dtype=dtype_dict)

            # Определяем колонки
            barcode_col = None
            sku_col = None
            destination_cols = []

            for i, col in enumerate(df.columns):
                col_str = str(col).lower().strip()
                if not col_str or 'итог' in col_str or 'sum' in col_str or col_str.startswith('='):
                    continue

                # Игнорируем первые две колонки (A=штрихкод, B=артикул) независимо от их названий
                if i < 2:
                    if 'штрихкод' in col_str or 'шк' in col_str or col_str == '0':
                        barcode_col = col
                    elif 'артикул' in col_str or 'арт' in col_str or col_str == '1':
                        sku_col = col
                    continue

                # Все остальные колонки - это направления
                destination_cols.append(col)

            if barcode_col is None and len(df.columns) > 0:
                barcode_col = df.columns[0]
            if sku_col is None and len(df.columns) > 1:
                sku_col = df.columns[1]

            if not destination_cols:
                destination_cols = [col for col in df.columns[2:] if str(col).strip() and not str(col).startswith('Unnamed')]

            valid_destinations = []
            for destination in destination_cols:
                has_data = False
                for _, row in df.iterrows():
                    if pd.notna(row[destination]) and str(row[destination]).strip():
                        try:
                            val = float(str(row[destination]))
                            if val > 0:
                                has_data = True
                                break
                        except (ValueError, TypeError):
                            continue
                if has_data:
                    valid_destinations.append(destination)

            destination_cols = valid_destinations

            if not destination_cols:
                QMessageBox.critical(
                    self.main_window, "Ошибка",
                    "Не найдены колонки с направлениями в файле!"
                )
                return

            # Сохраняем текущие коробки и их содержимое перед обновлением
            # Ключ: destination_name (без префикса группы)
            old_shipment_data = {}
            for shipment_name, shipment in group_shipment.sub_shipments.items():
                # Получаем чистое имя направления без префикса группы
                dest_name = getattr(shipment, 'display_name', None)
                if not dest_name:
                    # Извлекаем из original_destination_name или shipment_name
                    orig_name = getattr(shipment, 'original_destination_name', shipment_name)
                    if '::' in orig_name:
                        dest_name = orig_name.split('::', 1)[1]
                    else:
                        dest_name = orig_name

                old_allocated_qty = {}
                for barcode, item in shipment.shipment_items.items():
                    old_allocated_qty[barcode] = item.allocated_qty

                old_shipment_data[dest_name] = {
                    'shipment_object': shipment,  # Сохраняем ссылку на существующий объект
                    'boxes': shipment.boxes[:],
                    'current_box_index': shipment.current_box_index,
                    'removed_items': shipment.removed_items.copy(),
                    'partial_decrease_items': shipment.partial_decrease_items.copy(),
                    'properties': shipment.properties,
                    'allocated_qty': old_allocated_qty
                }

            # Сохраняем информацию о текущей поставке, если она принадлежит этой группе
            current_shipment_was_in_group = False
            current_shipment_dest_name = None
            if (self.main_window.current_shipment and
                self.main_window.current_shipment.parent_group and
                self.main_window.current_shipment.parent_group.group_name == group_shipment.group_name):
                current_shipment_was_in_group = True
                current_shipment_dest_name = getattr(self.main_window.current_shipment, 'display_name',
                    self.main_window.current_shipment.destination_name)

            # ИСПОЛЬЗУЕМ СУЩЕСТВУЮЩИЕ объекты Shipment вместо создания новых
            # Это сохраняет все ссылки на объекты валидными
            updated_sub_shipments = {}
            for destination in destination_cols:
                destination_name = str(destination).strip()
                if not destination_name:
                    continue

                # Проверяем, существует ли уже поставка с таким destination_name
                old_data = old_shipment_data.get(destination_name)

                if old_data:
                    # ИСПОЛЬЗУЕМ существующий объект Shipment - обновляем его на месте
                    shipment = old_data['shipment_object']
                    # Обновляем имена на случай если они не были установлены
                    shipment.display_name = destination_name
                    unique_shipment_name = f"{group_shipment.group_name}::{destination_name}"
                    shipment.original_destination_name = unique_shipment_name
                    # Очищаем старые товары
                    shipment.shipment_items.clear()
                    # Сбрасываем свойства коробки
                    shipment.boxes = old_data['boxes'][:]
                    shipment.current_box_index = old_data['current_box_index']
                    shipment.removed_items = old_data['removed_items'].copy()
                    shipment.partial_decrease_items = old_data['partial_decrease_items'].copy()
                    shipment.properties = old_data['properties']
                    shipment.parent_group = group_shipment
                else:
                    # Создаем новую поставку только если её не было в старой группе
                    unique_shipment_name = f"{group_shipment.group_name}::{destination_name}"
                    shipment = Shipment(
                        unique_shipment_name,
                        group_shipment.font_size,
                        group_shipment.label_font_size,
                        group_shipment.theme
                    )
                    shipment.display_name = destination_name
                    shipment.original_destination_name = unique_shipment_name
                    shipment.parent_group = group_shipment

                # Заполняем поставку данными из нового Excel
                for idx, row in df.iterrows():
                    barcode_raw = row[barcode_col] if pd.notna(row[barcode_col]) else ""
                    if barcode_raw == "" or pd.isna(barcode_raw):
                        continue

                    barcode = str(barcode_raw).strip()
                    sku = str(row[sku_col]).strip() if pd.notna(row[sku_col]) else ""

                    qty_value = row[destination]
                    if pd.isna(qty_value) or str(qty_value).strip() == "":
                        qty = 0
                    else:
                        try:
                            qty = int(float(str(qty_value)))
                        except (ValueError, TypeError):
                            qty = 0

                    if barcode and qty > 0:
                        if barcode in shipment.shipment_items:
                            shipment.shipment_items[barcode].total_qty += qty
                        else:
                            shipment.shipment_items[barcode] = ShipmentItem(barcode, sku, qty)

                # Пересчитываем allocated_qty на основе содержимого коробок
                shipment.recalculate_allocated_qty_from_boxes()
                shipment.invalidate_caches()

                if shipment.shipment_items:
                    updated_sub_shipments[destination_name] = shipment
                    self.main_window.logger.debug(f"Поставка {destination_name}: {len(shipment.shipment_items)} товаров, total_qty={sum(item.total_qty for item in shipment.shipment_items.values())}")

            if not updated_sub_shipments:
                QMessageBox.warning(
                    self.main_window, "Предупреждение",
                    "В файле не найдено данных для обновления поставок. Проверьте формат файла."
                )
                return

            # Очищаем кэш групповой поставки перед обновлением sub_shipments
            group_shipment.invalidate_caches()

            # Заменяем sub_shipments на обновлённые
            group_shipment.sub_shipments.clear()
            for dest_name, shipment in updated_sub_shipments.items():
                # Используем original_destination_name как ключ для избежания конфликтов
                shipment_key = getattr(shipment, 'original_destination_name', shipment.destination_name)
                group_shipment.sub_shipments[shipment_key] = shipment

                # Сохраняем поставку с коробками в БД
                self.main_window.shipment_manager.save_shipment(shipment, preserve_box_items=True)

            # Обрабатываем исчезнувшие товары - добавляем их в removed_items соответствующих поставок
            # Собираем все старые товары
            all_old_items = {}
            for dest_name, old_data in old_shipment_data.items():
                shipment = old_data['shipment_object']
                for barcode, item in shipment.shipment_items.items():
                    key = f"{dest_name}_{barcode}"
                    all_old_items[key] = {
                        'dest_name': dest_name,
                        'barcode': barcode,
                        'item': item,
                    }

            # Определяем какие товары исчезли (были в старых, но нет в новых)
            disappeared_items = set(all_old_items.keys())
            for dest_name, shipment in updated_sub_shipments.items():
                for barcode, item in shipment.shipment_items.items():
                    old_key = f"{dest_name}_{barcode}"
                    if old_key in disappeared_items:
                        disappeared_items.remove(old_key)

            # Добавляем исчезнувшие товары в removed_items
            for key in disappeared_items:
                item_info = all_old_items[key]
                dest_name = item_info['dest_name']
                barcode = item_info['barcode']
                old_item = item_info['item']

                # Находим соответствующую поставку в новом составе
                target_shipment = updated_sub_shipments.get(dest_name)

                if target_shipment:
                    # Проверяем, есть ли товар в коробках
                    allocated_from_boxes = sum(
                        box.items.get(barcode, 0) for box in target_shipment.boxes
                    )
                    if allocated_from_boxes > 0 and barcode not in target_shipment.shipment_items:
                        # Товар исчез из нового Excel, но остался в коробках
                        if barcode not in target_shipment.removed_items:
                            target_shipment.removed_items[barcode] = {
                                'sku': old_item.sku,
                                'allocated_qty': allocated_from_boxes
                            }
                        else:
                            target_shipment.removed_items[barcode]['allocated_qty'] += allocated_from_boxes
                        target_shipment.invalidate_caches()

            # Если текущая поставка была частью этой группы, обновляем ссылку на неё
            if current_shipment_was_in_group:
                found_shipment = updated_sub_shipments.get(current_shipment_dest_name)
                if found_shipment:
                    # current_shipment уже указывает на тот же объект (т.к. мы не создавали новые)
                    # Но на случай если была создана новая поставка, обновим ссылку
                    self.main_window.current_shipment = found_shipment

            # Очищаем кэш групповой поставки после обновления всех данных
            group_shipment.invalidate_caches()

            # Принудительно обновляем UI для всех связанных виджетов
            self.main_window.update_ui()
            utils.play_sound(self.main_window.ok_sound, self.main_window.tone_sound)

        except Exception as e:
            self.main_window.logger.error(f"Ошибка при обновлении групповой поставки: {e}", exc_info=True)
            # Rollback to original state if there was an error
            # Т.к. мы изменяли существующие объекты на месте, нужно восстановить их состояние
            for shipment_name, old_shipment in original_sub_shipments.items():
                # Восстанавливаем original_destination_name если он был изменён
                if '::' in shipment_name:
                    dest_name = shipment_name.split('::', 1)[1]
                    if hasattr(old_shipment, 'display_name'):
                        old_shipment.display_name = dest_name
                    old_shipment.original_destination_name = shipment_name

            group_shipment.sub_shipments = original_sub_shipments
            # Очищаем кэш после отката
            group_shipment.invalidate_caches()
            QMessageBox.critical(self.main_window, "Ошибка", f"Не удалось обновить групповую поставку:\n{str(e)}")

    def rename_group_shipment(self, old_name):
        existing_names = list(self.main_window.group_shipments.keys())
        dialog = RenameDialog(old_name, is_shipment=True, existing_names=existing_names, parent=self.main_window)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        new_name = dialog.get_new_name()
        if not new_name or new_name == old_name:
            return
        try:
            database.execute_query(
                "UPDATE shipments SET parent_group = %s WHERE parent_group = %s",
                (new_name, old_name)
            )
            group_shipment = self.main_window.group_shipments.pop(old_name)
            group_shipment.group_name = new_name
            self.main_window.group_shipments[new_name] = group_shipment
            self.main_window.update_ui()
            # Статус скрыт, сообщения не отображаются
            # self.main_window.statusBar().showMessage(f"Группа поставок переименована в «{new_name}»", 3000)
        except Exception as e:
            QMessageBox.critical(self.main_window, "Ошибка", f"Не удалось переименовать группу поставок:\n{e}")

    def delete_group_shipment(self, group_name):
        reply = QMessageBox.question(
            self.main_window, "Подтверждение",
            f"Вы уверены, что хотите удалить группу поставок «{group_name}» со всеми входящими в нее поставками?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        try:
            # ��ринудительно сохраняем все отложенные изменения перед удалением
            self.main_window.force_save_session()

            # Получаем все поставки в группе для их последующего удаления
            shipment_names = []
            group_shipment = self.main_window.group_shipments.get(group_name)
            if group_shipment:
                shipment_names = list(group_shipment.sub_shipments.keys())

                # Удаляем каждую поставку в группе с использованием правильной функции удаления
                for shipment_name in shipment_names:
                    # Используем функцию удаления архивированной поста��ки, так как она корректно удаляет все связанные данные
                    database.delete_archived_shipment(shipment_name)
                    
                    # Очищаем кэш для удалённой поставки
                    self.main_window.shipment_manager.clear_cache(shipment_name)

            # Удаляем гр��ппу из словаря
            del self.main_window.group_shipments[group_name]

            if (self.main_window.current_shipment and
                hasattr(self.main_window.current_shipment, 'parent_group') and
                self.main_window.current_shipment.parent_group and
                self.main_window.current_shipment.parent_group.group_name == group_name):
                self.main_window.current_shipment = None

            # Сбрасываем флаг updating_ui, чтобы обновление не блокировалось
            self.main_window.updating_ui = False
            
            # Принудительно обновляем дерево поставок для немедленного отображения изменений
            self.main_window.ui_updater.update_shipments_tree()
            # Статус скрыт, сообщения не о��ображаются
            # self.main_window.statusBar().showMessage(f"Группа поставок «{group_name}» удалена", 3000)
        except Exception as e:
            QMessageBox.critical(self.main_window, "Ошибка", f"Не удалось удалить группу поставок:\n{e}")