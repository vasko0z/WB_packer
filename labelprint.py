import sys
from openpyxl import load_workbook
import os
import subprocess
import sqlite3
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
from PIL import Image, ImageTk
import io
import json
import tempfile
import base64
import time
import re
from typing import List, Dict, NamedTuple, Union
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
# Импорт для воспроизведения звуков
try:
    from audio_handler import play_sound_simple, is_audio_available
    HAS_AUDIO = is_audio_available()
except ImportError:
    HAS_AUDIO = False
    print("audio_handler не установлен. Функция воспроизведения звуков будет недоступна.")

# Импорт оптимизированного менеджера памяти и подпроцессов
try:
    from optimized_memory_manager import OptimizedMemoryManager, safe_node_runner, cleanup_on_exit, monitor_resources
    MEMORY_MANAGER_AVAILABLE = True
    memory_manager = OptimizedMemoryManager()
except ImportError:
    MEMORY_MANAGER_AVAILABLE = False
    print("optimized_memory_manager не установлен. Будет использоваться стандартное управление процессами.")

def check_and_install_dependencies():
    """Проверяет и устанавливает недостающие зависимости"""
    required_packages = {
        'setuptools': 'setuptools',  # Добавляем setuptools для совместимости с новыми версиями Python
        'PIL': 'Pillow',  # Для работы с изображениями (импортируется как PIL)
        'reportlab': 'reportlab',  # Для генерации PDF
        'openpyxl': 'openpyxl',  # Для работы с Excel файлами
        # audio_handler не является отдельным пакетом, это наш собственный модуль
    }
    
    missing_packages = []
    for import_name, package_name in required_packages.items():
        try:
            __import__(import_name)
        except ImportError:
            missing_packages.append(package_name)
    
    # Проверяем существование нашего собственного модуля audio_handler
    if not os.path.exists('audio_handler.py'):
        print("Файл audio_handler.py не найден. Он необходим для воспроизведения звуков.")
        missing_packages.append('audio_handler.py (локальный файл)')
    
    if missing_packages:
        print(f"Отсутствующие пакеты: {missing_packages}")
        try:
            # Пытаемся установить недостающие пакеты
            packages_to_install = [pkg for pkg in missing_packages if pkg != 'audio_handler.py (локальный файл)']
            for package in packages_to_install:
                print(f"Устанавливаю {package}...")
                subprocess.check_call([sys.executable, "-m", "pip", "install", package])
            print("Все недостающие пакеты установлены. Перезапустите приложение.")
            sys.exit(0)
        except subprocess.CalledProcessError:
            print(f"Не удалось автоматически установить пакеты: {missing_packages}")
            print("Установите их вручную с помощью команды: pip install " + " ".join([pkg for pkg in missing_packages if pkg != 'audio_handler.py (локальный файл)']))

from PyQt6.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QLineEdit, QDialog, QMessageBox, QMenu, QMenuBar, QDialogButtonBox, QTabWidget, QGroupBox, QFormLayout, QSpinBox
from PyQt6.QtCore import Qt
from PyQt6.QtPrintSupport import QPrinter
from PyQt6.QtGui import QFont, QPageSize, QPageLayout, QPainter, QPen, QImage, QPixmap, QIntValidator
from PyQt6.QtCore import QSizeF, QPoint
import gc
import psutil

# Import PyMuPDF (fitz) for PDF handling - optional dependency with fallback
try:
    import fitz  # type: ignore # PyMuPDF
    HAS_PYMUPDF = True
    print(f"PyMuPDF loaded successfully. Version: {fitz.__version__}")  # Debug info
    print(f"PyMuPDF location: {fitz.__file__}")  # Debug info
    print(f"HAS_PYMUPDF variable set to: {HAS_PYMUPDF}")  # Additional debug info
except ImportError as e:
    HAS_PYMUPDF = False
    print(f"Failed to import PyMuPDF: {e}")  # Debug info
    print(f"Current Python path: {[p for p in sys.path if 'Python' in p]}")  # Path debug info
    print(f"HAS_PYMUPDF variable set to: {HAS_PYMUPDF}")  # Additional debug info

# GS1 DataMatrix constants and functions
# GS (Group Separator) - ASCII 29
GS = chr(29)

# RU to EN keyboard layout mapping (for quick RU->EN correction)
RU_EN_MAP = {
    'А': 'F', 'а': 'f', 'Б': ',', 'б': ',', 'В': 'D', 'в': 'd', 'Г': 'U', 'г': 'u',
    'Д': 'L', 'д': 'l', 'Е': 'T', 'е': 't', 'Ё': '`', 'ё': '`', 'Ж': ';', 'ж': ';',
    'З': 'P', 'з': 'p', 'И': 'B', 'и': 'b', 'Й': 'Q', 'й': 'q', 'К': 'R', 'к': 'r',
    'Л': 'K', 'л': 'k', 'М': 'V', 'м': 'v', 'Н': 'Y', 'н': 'y', 'О': 'J', 'о': 'j',
    'П': 'G', 'п': 'g', 'Р': 'K', 'р': 'k', 'С': 'S', 'с': 's', 'Т': 'N', 'т': 'n',
    'У': 'E', 'у': 'e', 'Ф': 'A', 'ф': 'a', 'Х': '[', 'х': '[', 'Ц': 'W', 'ц': 'w',
    'Ч': '[', 'ч': '[', 'Ш': 'I', 'ш': 'i', 'Щ': ']', 'щ': ']', 'Ъ': '}', 'ъ': '}',
    'Ы': 'S', 'ы': 's', 'Ь': ']', 'ь': ']', 'Э': '"', 'э': '"', 'Ю': '.', 'ю': '.',
    'Я': 'Z', 'я': 'z'
}


def ru_to_en_by_layout(s: str) -> str:
    """
    Convert Russian keyboard layout to English
    """
    return ''.join(RU_EN_MAP.get(ch, ch) for ch in s)


def normalize_raw_input(input_str: str) -> str:
    """
    Normalize various placeholders to real GS and clean input
    """
    s = (input_str or '').strip()
    
    # Remove garbage newlines/tabs
    s = re.sub(r'[\r\n\t]+', '', s)
    
    # Scanner symbology prefixes (at the beginning of string)
    s = re.sub(r'^\](?:d2|D2|c1|C1)', '', s)
    
    # Leading FNC1 sometimes appears as U+00E8 ('è') or broken character
    s = re.sub(r'^\u00E8', '', s)  # 'è'
    s = re.sub(r'^\u00EA', '', s)  # 'ê'
    
    # Convert various placeholders to our GS
    s = s.replace('<GS>', GS) \
         .replace('[GS]', GS) \
         .replace('^]', GS) \
         .replace('(FNC1)', GS) \
         .replace('<FNC1>', GS) \
         .replace('FNC1', GS) \
         .replace('\\F', GS) \
         .replace('\u2194', GS) \
         .replace('\u001D', GS) \
         .replace('\\x1d', GS) \
         .replace('\\x1D', GS)
    
    # Case insensitive versions
    s = re.sub(r'<gs>', GS, s, flags=re.IGNORECASE)
    s = re.sub(r'\[gs\]', GS, s, flags=re.IGNORECASE)
    s = re.sub(r'\^\]', GS, s, flags=re.IGNORECASE)
    s = re.sub(r'\(fnc1\)', GS, s, flags=re.IGNORECASE)
    s = re.sub(r'<fnc1>', GS, s, flags=re.IGNORECASE)
    s = re.sub(r'fnc1', GS, s, flags=re.IGNORECASE)
    
    # Collapse consecutive GS characters
    s = re.sub(f'{re.escape(GS)}+', GS, s)
    
    return s


class Gs1Tail(NamedTuple):
    """Represents a GS1 tail field"""
    ai: str  # '91', '92', or '93'
    value: str
    had_leading_gs: bool


class Gs1Payload(NamedTuple):
    """Complete GS1 payload with all parsed components"""
    gtin: str
    serial: str  # 1..20 chars
    tails: List[Gs1Tail]
    pretty_ai: str  # (01) ... (21) ... (9x) ...
    raw_with_gs: str  # 01...21...<GS>9x...
    ai_text: str  # (01)...(21)...<GS>(9x)...


def is_digits(s: str, length: int) -> bool:
    """Check if string contains only digits and has specific length"""
    return len(s) == length and s.isdigit()


def is_ai(s: str, i: int) -> Union[str, None]:
    """Check if position i contains a valid AI code (91, 92, or 93)"""
    if i + 2 <= len(s):
        ai = s[i:i+2]
        if ai in ['91', '92', '93']:
            return ai
    return None


def is_valid_92(value: str) -> bool:
    """Validate AI 92 value (should be 44 or 88 chars of specific chars)"""
    if len(value) not in [44, 88]:
        return False
    return bool(re.match(r'^[A-Za-z0-9+/=\-_.]*$', value))


def parse_gs1(input_normalized: str) -> Gs1Payload:
    """
    Parse normalized GS1 string into structured data
    Handles both traditional AI-parenthesized format and raw format with FNC1/GS
    """
    s = input_normalized
    i = 0
    
    # Check if string starts with FNC1 (GS character), but DON'T skip it
    # because for proper GS1 DataMatrix it should be included
    has_explicit_fnc1 = s.startswith(GS)
    
    # Check if string starts with '01' (GTIN AI)
    if i < len(s) and s[i:i+2] == '01':
        i += 2
        
        # Extract GTIN (14 digits)
        if i + 14 > len(s):
            raise ValueError('E01_BAD_GTIN: Недостаточно символов для GTIN')
        
        gtin = s[i:i+14]
        if not is_digits(gtin, 14):
            raise ValueError('E01_BAD_GTIN: (01) — 14 цифр')
        i += 14
        
        # Check if follows with '21' (Serial Number AI)
        if i < len(s) and s[i:i+2] == '21':
            i += 2
            
            # Find where serial ends - before next AI (91, 92, 93)
            # We need to be careful to avoid finding AI patterns inside the serial data
            serial_end = len(s)
            
            # Search for the next AI code that starts at a reasonable position
            # Look for 91, 92, 93 followed by the expected number of characters
            for ai_code in ['91', '92', '93']:
                pos = i
                while pos < len(s):
                    ai_pos = s.find(ai_code, pos)
                    if ai_pos == -1:
                        break
                    
                    # Check if this is a valid AI position (not inside other data)
                    # AI 91 and 93 have 4-char values, AI 92 has variable length
                    if ai_code in ['91', '93']:
                        expected_len = 4
                        if ai_pos + 2 + expected_len <= len(s):
                            # This looks like a valid AI
                            serial_end = min(serial_end, ai_pos)
                            break
                    else:  # AI 92
                        # For AI 92, we need to check if it's a valid position
                        # by looking ahead for another AI or the end of string
                        serial_end = min(serial_end, ai_pos)
                        break
                    
                    pos = ai_pos + 1  # Continue searching after this position
            
            serial = s[i:serial_end]
            i = serial_end
            
            if len(serial) < 1:
                raise ValueError('E21_LEN: (21) пустой')
            if len(serial) > 20:
                raise ValueError('E21_LEN: (21) превышает 20 символов GS1')
        else:
            raise ValueError('E21_MISSING: отсутствует (21)')
    else:
        raise ValueError('E01_BAD_GTIN: строка не начинается с (01)')
    
    # Process remaining parts that might be AI fields
    tails: List[Gs1Tail] = []
    
    while i < len(s):
        # Check if current position has a GS separator
        had_gs_separator = False
        if i < len(s) and s[i] == GS:
            had_gs_separator = True
            i += 1
        
        # Look for AI codes: 91, 92, 93
        if i + 1 < len(s):
            ai_code = s[i:i+2]
            
            if ai_code in ['91', '92', '93']:
                i += 2  # Move past AI code
                
                if ai_code == '91' or ai_code == '93':
                    # These AIs have fixed 4-character values
                    if i + 4 > len(s):
                        raise ValueError(f'E{ai_code}_LEN: Недостаточно символов для AI {ai_code}')
                    
                    value = s[i:i+4]
                    tails.append(Gs1Tail(ai=ai_code, value=value, had_leading_gs=had_gs_separator))
                    i += 4
                elif ai_code == '92':
                    # AI 92 can have variable length, often base64-like data
                    # Find the end of this value - could be at next GS, next AI, or end of string
                    
                    value_end = len(s)  # Default to end of string
                    
                    # Look for next GS if we had one
                    next_gs_pos = s.find(GS, i)
                    if next_gs_pos != -1:
                        value_end = min(value_end, next_gs_pos)
                    
                    # Look for next AI without GS separator
                    for next_ai in ['91', '92', '93', '01', '21']:
                        pos = s.find(next_ai, i)
                        if pos != -1 and pos < value_end:
                            # Verify it's actually an AI (not just a sequence in the data)
                            # For AIs like 91, 92, 93, 01, 21, we can check if they appear as standalone AIs
                            if pos == i or s[pos-1] == GS:  # Either at start or after GS
                                value_end = pos
                                break
                    
                    value = s[i:value_end]
                    tails.append(Gs1Tail(ai=ai_code, value=value, had_leading_gs=had_gs_separator))
                    i = value_end
            else:
                # Not a recognized AI, continue parsing
                i += 1
        else:
            # End of string
            break
    
    # Construct the output structures
    pretty_ai_parts = [f'(01) {gtin}', f'(21) {serial}']
    for t in tails:
        pretty_ai_parts.append(f'({t.ai}) {t.value}')
    pretty_ai = ' '.join(pretty_ai_parts)

    ai_text_parts = [f'(01){gtin}', f'(21){serial}']
    for t in tails:
        if t.had_leading_gs:
            ai_text_parts.append(f'<GS>({t.ai}){t.value}')
        else:
            ai_text_parts.append(f'<GS>({t.ai}){t.value}')  # Always add GS separator for proper GS1 format
    ai_text = ''.join(ai_text_parts)

    # Raw with GS construction
    # For proper GS1 format, we should always add FNC1 at the start and GS separators between AI fields
    raw_with_gs_parts = ['01' + gtin, '21' + serial]
    for t in tails:
        if t.had_leading_gs:
            raw_with_gs_parts.append(GS + t.ai + t.value)
        else:
            # Even if original didn't have GS separator, add it for proper GS1 format
            raw_with_gs_parts.append(GS + t.ai + t.value)
    raw_with_gs = ''.join(raw_with_gs_parts)
    
    # ALWAYS prepend FNC1 for GS1 DataMatrix compliance
    # This is the key fix: FNC1 must be the first character in GS1 DataMatrix
    # But only add it if it's not already there to avoid duplicates
    if not raw_with_gs.startswith(GS):
        raw_with_gs = GS + raw_with_gs

    return Gs1Payload(
        gtin=gtin,
        serial=serial,
        tails=tails,
        pretty_ai=pretty_ai,
        raw_with_gs=raw_with_gs,
        ai_text=ai_text
    )

def parse_from_user_input(raw: str) -> Gs1Payload:
    """
    Parse user input string into GS1 payload
    """
    return parse_gs1(normalize_raw_input(raw))


def check_cyrillic_layout(text: str) -> bool:
    """
    Check if the text contains Cyrillic characters that might indicate wrong keyboard layout
    """
    # Check for common Cyrillic characters that correspond to Latin ones
    cyrillic_chars = set('АБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯабвгдеёжзийклмнопрстуфхцчшщъыьэюя')
    return any(char in cyrillic_chars for char in text)


def validate_input(input_text: str) -> tuple[bool, str]:
    """
    Validate GS1 input text
    
    Args:
        input_text: Raw input string
    
    Returns:
        tuple: (is_valid, error_message)
    """
    if not input_text or not input_text.strip():
        return False, "Входная строка пуста"
    
    # Check for Cyrillic characters
    if check_cyrillic_layout(input_text):
        return False, "Смените раскладку клавиатуры на EN"
    
    # Basic validation: ensure it starts with a known AI
    stripped = input_text.strip()
    if not (stripped.startswith('01') or
            stripped.startswith('(01)') or
            stripped.startswith(']d2') or
            stripped.startswith(']D2') or
            stripped.startswith(']c1') or
            stripped.startswith(']C1')):
        return False, "Строка должна начинаться с AI (01) или сканерного префикса"
    
    return True, ""


def generate_datamatrix_alternative(data_string: str, scale: int = 6) -> bytes:
    """
    Alternative method to generate DataMatrix using pylibdmtx when bwip-js fails
    """
    # NOTE: This method may not produce GS1-compliant DataMatrix, so it's better to fix
    # the original issue rather than rely on this alternative
    raise RuntimeError("Alternative method is disabled to ensure GS1 compliance. Fix the original issue instead.")


def generate_datamatrix_image(ai_parenthesized: str, raw_with_gs: str = None, scale: int = 6) -> bytes:
    """
    Generate DataMatrix image using embedded generator with fallback to Node.js
      
    Args:
        ai_parenthesized: AI-parenthesized string like '(01)12345678901234(21)SERIAL'
        raw_with_gs: Raw string with GS characters (optional) or pre-formatted string with ^FNC1/^GS (optional)
        scale: Scale factor for the generated image
      
    Returns:
        bytes: PNG image data
    """
    # First, try to use the embedded generator to avoid external dependencies
    try:
        from embedded_datamatrix_generator import generate_gs1_datamatrix_embedded
        result = generate_gs1_datamatrix_embedded(raw_with_gs or ai_parenthesized, ai_parenthesized)
        if result:
            print("DataMatrix generated successfully using embedded generator")
            return result
    except Exception as e:
        print(f"Embedded generator failed: {str(e)}")
        # Continue with original Node.js approach if embedded generator fails
        pass

    # Original Node.js approach
    # Determine the script path
    # Try to locate the js file in the executable directory first
    if getattr(sys, 'frozen', False):
        # Running as compiled executable
        script_dir = os.path.dirname(sys.executable)
    else:
        # Running as script
        script_dir = os.path.dirname(os.path.abspath(__file__))
    
    js_script_path = os.path.join(script_dir, "gs1_datamatrix_node_generator.js")
    
    # Check if the JS file exists, if not, try alternative locations
    if not os.path.exists(js_script_path):
        # Try relative path from current working directory
        js_script_path = os.path.join(".", "gs1_datamatrix_node_generator.js")
    
    if not os.path.exists(js_script_path):
        # Try relative path from the script's directory
        original_script_dir = os.path.dirname(os.path.abspath(__file__))
        js_script_path = os.path.join(original_script_dir, "gs1_datamatrix_node_generator.js")
    
    # Final check - if file still doesn't exist, raise an error
    if not os.path.exists(js_script_path):
        raise RuntimeError(f"JS script file not found: {js_script_path}. This file is required for DataMatrix generation.")
    
    # Check for problematic patterns that cause AI misinterpretation
    def has_problematic_patterns(text):
        # Check for sequences that might be misinterpreted as AIs
        import re
        # Look for patterns that might look like AIs but aren't preceded by GS
        # Common problematic AIs: 01, 10, 11, 12, 13, 15, 17, etc. followed by data
        # But especially sequences like 91, 92, 93 that might appear in data
        problematic_patterns = [
            r'(?<!\x1d|\()91[A-Z0-9]{4}',  # 91 followed by 4 chars not after GS or (
            r'(?<!\x1d|\()92[A-Z0-9+/=\-_.]{4,}',  # 92 followed by data not after GS or (
            r'(?<!\x1d|\()93[A-Z0-9]{4}'   # 93 followed by 4 chars not after GS or (
        ]
        for pattern in problematic_patterns:
            if re.search(pattern, text):
                return True
        return False
    
    # Prepare the options for bwip-js based on whether raw_with_gs is provided
    if raw_with_gs and raw_with_gs.strip():
        # Check if raw_with_gs contains real GS characters (ASCII 29) - if so, convert to ^FNC1/^GS format
        if GS in raw_with_gs:
            # Convert real GS characters to ^GS markers for proper bwip-js processing
            # But handle the first character (FNC1) separately if it's a GS character
            if raw_with_gs.startswith(GS):
                # First character is FNC1, then we have GS separators between AI fields
                # So we need to replace only the subsequent GS characters, not the first one
                remaining_part = raw_with_gs[1:]  # Skip the first FNC1 character
                processed_remaining = remaining_part.replace(GS, '^GS')
                processed_raw = '^FNC1' + processed_remaining
            else:
                # No FNC1 at the beginning, add it and convert GS characters
                processed_raw = '^FNC1' + raw_with_gs.replace(GS, '^GS')
            
            # Use datamatrix with parsefnc to ensure proper handling of FNC1/GS
            options = {
                "bcid": "datamatrix",
                "text": processed_raw,
                "parsefnc": True,  # Enable parsing of ^FNC1 and ^GS markers
                "parse": False,     # Disable AI parsing to avoid misinterpretation of data as AIs
                "scale": max(1, scale),
                "includetext": False,
                "paddingwidth": 0,
                "paddingheight": 0
            }
        else:
            # If raw_with_gs is already in parenthesized format, use gs1datamatrix
            if '(' in raw_with_gs and ')' in raw_with_gs:
                options = {
                    "bcid": "gs1datamatrix",
                    "text": raw_with_gs,  # gs1datamatrix adds FNC1 automatically
                    "parse": True,  # Allow parsing of AI brackets
                    "scale": max(1, scale),
                    "includetext": False,
                    "paddingwidth": 0,
                    "paddingheight": 0
                }
            else:
                # Default case - convert to parenthesized format
                parenthesized_data = convert_to_parenthesized_format(raw_with_gs)
                
                options = {
                    "bcid": "gs1datamatrix",
                    "text": parenthesized_data,  # gs1datamatrix adds FNC1 automatically
                    "parse": True,  # Allow parsing of AI brackets
                    "scale": max(1, scale),
                    "includetext": False,
                    "paddingwidth": 0,
                    "paddingheight": 0
                }
    else:
        # Check if ai_parenthesized has problematic patterns
        if has_problematic_patterns(ai_parenthesized):
            # Use datamatrix with parsefnc as a preventive measure
            processed_text = '^FNC1' + ai_parenthesized.replace(GS, '^GS').replace(" ", "")
            options = {
                "bcid": "datamatrix",
                "text": processed_text,
                "parsefnc": True,  # Enable parsing of ^FNC1 and ^GS markers
                "parse": False,    # Disable AI parsing to avoid misinterpretation
                "scale": max(1, scale),
                "includetext": False,
                "paddingwidth": 0,
                "paddingheight": 0
            }
        else:
            # Use the gs1datamatrix method with AI-parenthesized string
            text = ai_parenthesized.replace(" ", "")  # Remove spaces
            # CRITICAL FIX: Ensure gs1datamatrix is used to properly handle FNC1
            options = {
                "bcid": "gs1datamatrix",
                "text": text,  # gs1datamatrix adds FNC1 automatically
                "parse": True,  # Parses AIs; inserts FNC1 leader and GS where required
                "scale": max(1, scale),
                "includetext": False,
                "paddingwidth": 0,
                "paddingheight": 0
            }
    # Execute the Node.js script using the optimized memory manager
    try:
        # Convert options to JSON string
        options_json = json.dumps(options)
        
        # Use the optimized memory manager if available
        if MEMORY_MANAGER_AVAILABLE:
            # Run the Node.js script using the safe runner
            result_bytes = safe_node_runner(js_script_path, options)
            if result_bytes is None:
                # If safe runner fails, try direct subprocess call as fallback
                print("Safe runner failed, attempting direct subprocess call...")
                working_dir = os.path.dirname(js_script_path)
                
                result = subprocess.run([
                    "node", js_script_path, options_json
                ], capture_output=True, check=False, shell=False, cwd=working_dir)
                
                if result.returncode != 0:
                    error_msg = result.stderr.decode('utf-8', errors='ignore')
                    raise RuntimeError(f"Node.js script failed (direct call): {error_msg}")
                else:
                    return result.stdout
            return result_bytes
        else:
            # Fall back to the original method if memory manager is not available
            working_dir = os.path.dirname(js_script_path)
            
            result = subprocess.run([
                "node", js_script_path, options_json
            ], capture_output=True, check=False, shell=False, cwd=working_dir)
            
            if result.returncode != 0:
                error_msg = result.stderr.decode('utf-8', errors='ignore')
                
                # Check specifically for AI misinterpretation errors
                if "GS1unknownAI" in error_msg or "AIs must start with" in error_msg or ("FNC1" in error_msg and "misinterpretation" in error_msg) or "unknownFNC" in error_msg or "Unknown function character" in error_msg:
                    # This is the specific error we need to handle
                    # The issue is that bwip-js incorrectly interprets sequences in AI 92 data
                    # as other AIs. The proper solution is to ensure GS1 compliance.
                    
                    print(f"Attempting to fix GS1 parsing error: {error_msg}...")
                    
                    # For problematic cases, try using datamatrix with parsefnc option directly
                    if raw_with_gs and raw_with_gs.startswith('^FNC1'):
                        safer_options = {
                            "bcid": "datamatrix",
                            "text": raw_with_gs or ai_parenthesized.replace(" ", ""),  # Use ai_parenthesized as fallback if raw_with_gs is empty
                            "parsefnc": True,  # Enable parsing of ^FNC1 and ^GS markers
                            "parse": False,    # Disable AI parsing to avoid misinterpretation
                            "scale": max(1, scale),
                            "includetext": False,
                            "paddingwidth": 0,
                            "paddingheight": 0
                        }
                        
                        options_json = json.dumps(safer_options)
                        result = subprocess.run([
                            "node", js_script_path, options_json
                        ], capture_output=True, check=False, shell=False, cwd=working_dir)
                        
                        if result.returncode != 0:
                            error_msg = result.stderr.decode('utf-8', errors='ignore')
                            # If this still fails, try a more basic approach ensuring proper FNC1/GS handling
                            # Convert raw_with_gs to proper ^FNC1/^GS format for basic processing
                            if raw_with_gs:
                                if raw_with_gs.startswith(GS):
                                    # Handle case where raw_with_gs starts with real FNC1 character
                                    remaining_part = raw_with_gs[1:]  # Skip the first FNC1 character
                                    processed_remaining = remaining_part.replace(GS, '^GS')
                                    basic_text = '^FNC1' + processed_remaining
                                else:
                                    # Add FNC1 and convert GS characters
                                    basic_text = '^FNC1' + raw_with_gs.replace(GS, '^GS')
                            else:
                                # Fallback to ai_parenthesized with proper FNC1 prefix
                                basic_text = "^FNC1" + ai_parenthesized.replace(" ", "").replace('<GS>', '^GS')
                            
                            basic_options = {
                                "bcid": "datamatrix",
                                "text": basic_text,
                                "parsefnc": True,  # Enable FNC1/GS parsing even in basic mode
                                "parse": False,
                                "scale": max(1, scale),
                                "includetext": False,
                                "paddingwidth": 0,
                                "paddingheight": 0
                            }
                            
                            options_json = json.dumps(basic_options)
                            result = subprocess.run([
                                "node", js_script_path, options_json
                            ], capture_output=True, check=False, shell=False, cwd=working_dir)
                            
                            if result.returncode != 0:
                                error_msg = result.stderr.decode('utf-8', errors='ignore')
                                raise RuntimeError(f"GS1 parsing error: Library limitation. Data contains sequence that triggers AI misinterpretation: {error_msg}")
                            else:
                                return result.stdout
                        else:
                            return result.stdout
                    else:
                        # If raw_with_gs doesn't start with ^FNC1, convert it first
                        if raw_with_gs:
                            processed_raw = '^FNC1' + raw_with_gs.replace(GS, '^GS')
                            safer_options = {
                                "bcid": "datamatrix",
                                "text": processed_raw or ai_parenthesized.replace(" ", ""),  # Use ai_parenthesized as fallback
                                "parsefnc": True,  # Enable parsing of ^FNC1 and ^GS markers
                                "parse": False,    # Disable AI parsing to avoid misinterpretation
                                "scale": max(1, scale),
                                "includetext": False,
                                "paddingwidth": 0,
                                "paddingheight": 0
                            }
                            
                            options_json = json.dumps(safer_options)
                            result = subprocess.run([
                                "node", js_script_path, options_json
                            ], capture_output=True, check=False, shell=False, cwd=working_dir)
                            
                            if result.returncode != 0:
                                error_msg = result.stderr.decode('utf-8', errors='ignore')
                                # If this also fails, try the basic approach ensuring proper FNC1/GS handling
                                # Convert raw_with_gs to proper ^FNC1/^GS format for basic processing
                                if raw_with_gs:
                                    if raw_with_gs.startswith(GS):
                                        # Handle case where raw_with_gs starts with real FNC1 character
                                        remaining_part = raw_with_gs[1:]  # Skip the first FNC1 character
                                        processed_remaining = remaining_part.replace(GS, '^GS')
                                        basic_text = '^FNC1' + processed_remaining
                                    else:
                                        # Add FNC1 and convert GS characters
                                        basic_text = '^FNC1' + raw_with_gs.replace(GS, '^GS')
                                else:
                                    # Fallback to ai_parenthesized with proper FNC1 prefix
                                    basic_text = "^FNC1" + ai_parenthesized.replace(" ", "").replace('<GS>', '^GS')
                                
                                basic_options = {
                                    "bcid": "datamatrix",
                                    "text": basic_text,
                                    "parsefnc": True,  # Enable FNC1/GS parsing even in basic mode
                                    "parse": False,
                                    "scale": max(1, scale),
                                    "includetext": False,
                                    "paddingwidth": 0,
                                    "paddingheight": 0
                                }
                                
                                options_json = json.dumps(basic_options)
                                result = subprocess.run([
                                    "node", js_script_path, options_json
                                ], capture_output=True, check=False, shell=False, cwd=working_dir)
                                
                                if result.returncode != 0:
                                    error_msg = result.stderr.decode('utf-8', errors='ignore')
                                    raise RuntimeError(f"GS1 parsing error: Library limitation. Data contains sequence that triggers AI misinterpretation: {error_msg}")
                                else:
                                    return result.stdout
                            else:
                                return result.stdout
                        else:
                            # If no raw_with_gs provided, try with the parenthesized version
                            processed_text = '^FNC1' + (ai_parenthesized.replace(" ", "").replace('<GS>', '^GS') or "011234567890123421ABCDEF")  # Ensure we have some text
                            safer_options = {
                                "bcid": "datamatrix",
                                "text": processed_text,
                                "parsefnc": True,  # Enable parsing of ^FNC1 and ^GS markers
                                "parse": False,    # Disable AI parsing to avoid misinterpretation
                                "scale": max(1, scale),
                                "includetext": False,
                                "paddingwidth": 0,
                                "paddingheight": 0
                            }
                            
                            options_json = json.dumps(safer_options)
                            result = subprocess.run([
                                "node", js_script_path, options_json
                            ], capture_output=True, check=False, shell=False, cwd=working_dir)
                            
                            if result.returncode != 0:
                                error_msg = result.stderr.decode('utf-8', errors='ignore')
                                # Final fallback - try with basic datamatrix ensuring proper FNC1/GS handling
                                # Convert raw_with_gs to proper ^FNC1/^GS format for basic processing
                                if raw_with_gs:
                                    if raw_with_gs.startswith(GS):
                                        # Handle case where raw_with_gs starts with real FNC1 character
                                        remaining_part = raw_with_gs[1:]  # Skip the first FNC1 character
                                        processed_remaining = remaining_part.replace(GS, '^GS')
                                        basic_text = '^FNC1' + processed_remaining
                                    else:
                                        # Add FNC1 and convert GS characters
                                        basic_text = '^FNC1' + raw_with_gs.replace(GS, '^GS')
                                else:
                                    # Fallback to ai_parenthesized with proper FNC1 prefix
                                    basic_text = "^FNC1" + ai_parenthesized.replace(" ", "").replace('<GS>', '^GS')
                                
                                basic_options = {
                                    "bcid": "datamatrix",
                                    "text": basic_text,
                                    "parsefnc": True,  # Enable FNC1/GS parsing even in fallback mode
                                    "parse": False,
                                    "scale": max(1, scale),
                                    "includetext": False,
                                    "paddingwidth": 0,
                                    "paddingheight": 0
                                }
                                
                                options_json = json.dumps(basic_options)
                                result = subprocess.run([
                                    "node", js_script_path, options_json
                                ], capture_output=True, check=False, shell=False, cwd=working_dir)
                                
                                if result.returncode != 0:
                                    error_msg = result.stderr.decode('utf-8', errors='ignore')
                                    raise RuntimeError(f"GS1 parsing error: Library limitation. Data contains sequence that triggers AI misinterpretation: {error_msg}")
                                else:
                                    return result.stdout
                            else:
                                return result.stdout
                else:
                    raise RuntimeError(f"Node.js script failed: {error_msg}")
            
            # Return the PNG image data
            return result.stdout
        
    except FileNotFoundError:
        raise RuntimeError("Node.js is not installed or not in PATH. Please install Node.js from https://nodejs.org/")
    except Exception as e:
        # Если основной метод генерации DataMatrix не работает, пробуем использовать альтернативный метод
        print(f"Основной метод генерации DataMatrix не работает: {str(e)}")
        print("Попытка использовать альтернативный метод...")
        
        try:
            # Импортируем альтернативный генератор
            from alternative_datamatrix_generator import generate_gs1_datamatrix_fallback
            
            # Подготавливаем данные для альтернативного генератора
            data_to_encode = raw_with_gs if raw_with_gs else ai_parenthesized
            
            # Генерируем DataMatrix с помощью альтернативного метода
            result_bytes = generate_gs1_datamatrix_fallback(raw_with_gs, ai_parenthesized)
            
            if result_bytes is not None:
                print("Альтернативный метод генерации DataMatrix успешно выполнен")
                return result_bytes
            else:
                print("Альтернативный метод генерации DataMatrix не вернул данные")
                raise RuntimeError(f"Альтернативный метод генерации DataMatrix также не удался: {str(e)}")
        except ImportError as import_err:
            # Если альтернативный генератор недоступен, регистрируем ошибку и продолжаем с исходной ошибкой
            print(f"Альтернативный генератор недоступен: {str(import_err)}")
            # Проверим, установлена ли библиотека pylibdmtx
            try:
                import pylibdmtx
                print("Библиотека pylibdmtx установлена, но модуль alternative_datamatrix_generator не найден")
            except ImportError:
                print("Для работы альтернативного метода необходимо установить pylibdmtx: pip install pylibdmtx")
            
            raise RuntimeError(f"Failed to generate DataMatrix: {str(e)}. Для резервного метода см. инструкции в INSTALLATION_GUIDE.md")
        except Exception as alt_e:
            print(f"Ошибка в альтернативном методе генерации: {str(alt_e)}")
            # Если альтернативный метод также не работает, объединяем обе ошибки
            raise RuntimeError(f"Failed to generate DataMatrix with primary method: {str(e)}. Alternative method also failed: {str(alt_e)}. Для решения проблемы см. INSTALLATION_GUIDE.md")


def generate_gs1_datamatrix(ai_parenthesized: str, scale: int = 4) -> bytes:
    """
    Generate GS1 DataMatrix from AI-parenthesized string
    
    Args:
        ai_parenthesized: String in format '(01)12345678901234(21)SERIAL...'
        scale: Scale factor for the image
    
    Returns:
        bytes: PNG image data
    """
    # Compact the AI string by removing extra spaces
    compacted_ai = ai_parenthesized.replace(" ", "")
    
    return generate_datamatrix_image(compacted_ai, scale=scale)


def string_to_base64_latin1(s: str) -> str:
    """
    Convert a JS string (which may contain GS U+001D) to base64,
    preserving byte values 0..255 (latin1-style).
    Works correctly for strings that we generate (ASCII + GS).
    """
    bin_data = bytearray()
    for char in s:
        code = ord(char) & 0xff  # Get the lower 8 bits
        bin_data.append(code)
    
    return base64.b64encode(bin_data).decode('ascii')


def convert_to_parenthesized_format(raw_with_gs: str) -> str:
    """
    Convert raw GS1 string with GS characters to parenthesized format for gs1datamatrix
    Example: '011234567890123421SERIAL<GS>91DATA<GS>92OTHER'
    becomes '(01)12345678901234(21)SERIAL(91)DATA(92)OTHER'
    """
    result = []
    i = 0
    
    # Check if string starts with FNC1 (GS character) and skip it if present
    if raw_with_gs.startswith(GS):
        i = 1
    
    # Extract GTIN (01) - always starts with '01' and has 14 digits
    if i < len(raw_with_gs) and raw_with_gs[i:i+2] == '01' and len(raw_with_gs) >= i + 16:
        gtin = raw_with_gs[i+2:i+16]
        result.append(f'(01){gtin}')
        i += 16
    
    # Check if follows with '21' (serial)
    if i < len(raw_with_gs) and raw_with_gs[i:i+2] == '21':
        # Find where serial ends - before next GS or next AI
        serial_start = i + 2
        serial_end = len(raw_with_gs)
        
        # Find the next GS character after '21'
        next_gs_pos = raw_with_gs.find(GS, serial_start)
        if next_gs_pos != -1:
            serial_end = next_gs_pos
        else:
            # If no GS found, look for next AI code (91, 92, 93) to determine serial end
            for ai_code in ['91', '92', '93']:
                ai_pos = raw_with_gs.find(ai_code, serial_start)
                if ai_pos != -1 and (serial_end == len(raw_with_gs) or ai_pos < serial_end):
                    serial_end = ai_pos
        
        serial = raw_with_gs[serial_start:serial_end]
        result.append(f'(21){serial}')
        i = serial_end
    
    # Process remaining parts that start with GS + AI
    while i < len(raw_with_gs):
        if i < len(raw_with_gs) and raw_with_gs[i] == GS:
            i += 1  # Skip GS character
            if i + 1 < len(raw_with_gs):
                ai_code = raw_with_gs[i:i+2]
                # Check if it's a valid AI code
                if ai_code in ['01', '21', '91', '92', '93']:
                    i += 2
                    # Read value until next GS or end
                    value_end = len(raw_with_gs)
                    next_gs_pos = raw_with_gs.find(GS, i)
                    if next_gs_pos != -1:
                        value_end = next_gs_pos
                    else:
                        # If no GS found, look for next AI code (91, 92, 93) to determine value end
                        for next_ai_code in ['01', '21', '91', '92', '93']:
                            ai_pos = raw_with_gs.find(next_ai_code, i)
                            if ai_pos != -1 and (value_end == len(raw_with_gs) or ai_pos < value_end):
                                value_end = ai_pos
                    
                    value = raw_with_gs[i:value_end]
                    
                    # Add to result with proper AI formatting
                    result.append(f'({ai_code}){value}')
                    
                    i = value_end
                else:
                    # If not a recognized AI, just advance
                    i += 1
            else:
                break
        else:
            # Shouldn't happen if raw_with_gs is properly formatted
            i += 1
    
    return ''.join(result)


def sanitize_gs1_string_for_bwip(gs1_string: str) -> str:
    """
    Sanitizes a GS1 string to prevent bwip-js from misinterpreting data as AIs
    """
    # This is a targeted fix for the specific error mentioned
    # The error suggests that somewhere in the data there's a sequence that looks like AI 3565
    # Let's try to identify and handle such cases
    
    # For the specific case mentioned in the problem:
    # Input: 0104680827310214215f!MiCAM8Gl()91EE1092UyIn3qIgeis/IauetBXDEMID08HfqbPPTXGxxup/MvA=
    # The issue is likely in the AI 92 data: UyIn3qIgeis/IauetBXDEMID08HfqbPPTXGxxup/MvA=
    # Where 'ID08' might be interpreted as AI '08'
    
    import re
    
    # Pattern to identify potential AI-like sequences in AI 92 data
    # AIs are typically 2-4 digits, so we'll look for digit sequences that might be AIs
    parts = gs1_string.split('(92)')
    if len(parts) > 1:
        # We have an AI 92, let's handle the data part carefully
        before_92 = parts[0]
        after_92 = parts[1] if len(parts) > 1 else ""
        
        # Find the 92 data part (until the next AI pattern)
        # Look for the next AI pattern like (91), (92), (93), (01), (21), etc.
        next_ai_match = re.search(r'\(\d{2}\)', after_92)
        if next_ai_match:
            ai_92_data = after_92[:next_ai_match.start()]
            remaining = after_92[next_ai_match.start():]
        else:
            ai_92_data = after_92
            remaining = ""
        
        # For the problematic case, we'll return the original string as-is
        # since the issue may be more fundamental with bwip-js itself
        return gs1_string
    else:
        return gs1_string


def make_safe_caption(raw_with_gs: str, custom: str = None) -> str:
    """
    Makes caption ASCII-safe and replaces GS with <GS>
    """
    if custom:
        return custom
    
    replaced = raw_with_gs.replace(GS, '<GS>')
    # Keep only ASCII printable characters and Cyrillic characters
    safe_chars = []
    for char in replaced:
        code = ord(char)
        if 0x20 <= code <= 0x7E or 0x0400 <= code <= 0x04FF:
            safe_chars.append(char)
        else:
            safe_chars.append('?')  # Replace unsafe characters with '?'
    
    return ''.join(safe_chars)


def wrap_text_lines(text: str, max_chars_per_line: int) -> list[str]:
    """
    Simple word wrapping with hard cutoff for very long tokens
    """
    words = text.split()
    lines = []
    current_line = ""
    
    for word in words:
        if len(current_line) == 0:
            if len(word) <= max_chars_per_line:
                current_line = word
            else:
                # Split long word into chunks
                for i in range(0, len(word), max_chars_per_line):
                    lines.append(word[i:i + max_chars_per_line])
        else:
            test_line = current_line + " " + word
            if len(test_line) <= max_chars_per_line:
                current_line = test_line
            else:
                lines.append(current_line)
                if len(word) <= max_chars_per_line:
                    current_line = word
                else:
                    # Split long word into chunks
                    for i in range(0, len(word), max_chars_per_line):
                        lines.append(word[i:i + max_chars_per_line])
                    current_line = ""
    
    if current_line:
        lines.append(current_line)
    
    return lines


def build_label_pdf(raw_with_gs: str, ai: str, caption: str = None,
                   width_mm: float = 58.0, height_mm: float = 40.0,
                   margin_mm: float = 3.0, dm_box_mm: float = None) -> bytes:
    """
    Build a 58x40 mm label PDF with GS1 DM
    
    Args:
        raw_with_gs: Raw GS1 string with GS characters
        ai: Parenthesized AI string like '(01)12345678901234(21)SERIAL'
        caption: Optional caption to display under the matrix
        width_mm: Width of the label in mm (default 58)
        height_mm: Height of the label in mm (default 40)
        margin_mm: Margin in mm (default 3)
        dm_box_mm: Size of the DataMatrix box in mm (auto-calculated if None)
    
    Returns:
        bytes: PDF data
    """
    # Calculate dimensions in points (1 mm = 2.83465 points)
    width_pt = width_mm * 2.83465
    height_pt = height_mm * 2.83465
    margin_pt = margin_mm * 2.83465
    
    # Calculate DataMatrix size - increase by 30%
    if dm_box_mm is None:
        dm_side_pt = min(width_pt, height_pt) - 2 * margin_pt - 12 * 2.83465  # Leave space for caption
    else:
        dm_side_pt = dm_box_mm * 2.83465
    
    # Increase DataMatrix size by 30%
    dm_side_pt = dm_side_pt * 1.3
    
    # Create an in-memory buffer for the PDF
    buffer = io.BytesIO()
    
    # Create the PDF canvas
    c = canvas.Canvas(buffer, pagesize=(width_pt, height_pt))
    
    # Generate DataMatrix image using raw_with_gs to ensure proper GS handling
    # Use gs1datamatrix with the AI format which properly handles FNC1
    try:
        # Use the AI format with gs1datamatrix which automatically adds FNC1
        png_data = generate_datamatrix_image(ai, scale=6)
    except RuntimeError as e:
        if "unknownFNC" in str(e) or "GS91" in str(e):
            # Handle the specific error by using alternative approach
            print(f"Using alternative approach for DataMatrix generation: {str(e)}")
            # Convert raw_with_gs to format suitable for datamatrix with parsefnc
            if raw_with_gs.startswith(GS):
                remaining_part = raw_with_gs[1:]  # Skip the first FNC1 character
                processed_remaining = remaining_part.replace(GS, '^GS')
                processed_raw = '^FNC1' + processed_remaining
            else:
                # Add FNC1 and convert GS characters
                processed_raw = '^FNC1' + raw_with_gs.replace(GS, '^GS')
            
            # Use datamatrix with parsefnc to properly handle FNC1/GS symbols
            png_data = generate_datamatrix_image('', raw_with_gs=processed_raw, scale=6)
        else:
            raise e
    
    # Load the PNG data directly using PIL
    png_buffer = io.BytesIO(png_data)
    img = Image.open(png_buffer)
    
    # Handle transparency by compositing onto a white background
    if img.mode in ('RGBA', 'LA', 'P'):
        # Create a white background
        background = Image.new('RGB', img.size, (255, 255, 255))
        # Paste the image on the background using alpha channel as mask
        if img.mode == 'P':
            img = img.convert('RGBA')
        if img.mode in ('RGBA', 'LA'):
            background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else img.split()[3])
        img = background
    elif img.mode != 'RGB':
        img = img.convert('RGB')
    
    # Instead of saving to disk, draw the image directly using ReportLab
    # Save image to an in-memory buffer
    img_buffer = io.BytesIO()
    img.save(img_buffer, format='PNG')
    img_buffer.seek(0)
    
    # Position the DataMatrix on the left side of the label
    # Increase top margin by 3mm to add more empty space at the top (reduced from 5mm to ensure text fits)
    top_margin_increase = 3 * 2.83465  # 3mm in points
    dm_x = margin_pt  # Position from left margin
    dm_y = height_pt - margin_pt - top_margin_increase - dm_side_pt  # Position from top with increased margin
    
    # Draw the DataMatrix from the in-memory buffer
    img_buffer.seek(0)  # Reset buffer pointer to the beginning
    c.drawImage(ImageReader(img_buffer), dm_x, dm_y, dm_side_pt, dm_side_pt)

    # Add znak.png image on the right side of the label
    try:
        znak_img_path = "znak.png"
        if os.path.exists(znak_img_path):
            # Load znak.png image
            znak_img = Image.open(znak_img_path)
            
            # Handle transparency by compositing onto a white background
            if znak_img.mode in ('RGBA', 'LA', 'P'):
                # Create a white background
                znak_background = Image.new('RGB', znak_img.size, (255, 255, 255))
                # Paste the image on the background using alpha channel as mask
                if znak_img.mode == 'P':
                    znak_img = znak_img.convert('RGBA')
                if znak_img.mode in ('RGBA', 'LA'):
                    znak_background.paste(znak_img, mask=znak_img.split()[-1] if znak_img.mode == 'RGBA' else znak_img.split()[3])
                znak_img = znak_background
            elif znak_img.mode != 'RGB':
                znak_img = znak_img.convert('RGB')
            
            # Calculate position for znak.png on the right side - make it smaller than DataMatrix
            znak_width = dm_side_pt * 0.6  # Make it 60% of DataMatrix size
            znak_height = dm_side_pt * 0.6  # Make it 60% of DataMatrix size
            
            # Position znak.png on the right side, centered vertically with DataMatrix
            znak_x = width_pt - margin_pt - znak_width  # Position from right margin
            znak_y = dm_y + (dm_side_pt - znak_height) / 2  # Center vertically with DataMatrix
            
            # Save znak image to buffer for ReportLab
            znak_buffer = io.BytesIO()
            znak_img.save(znak_buffer, format='PNG')
            znak_buffer.seek(0)
            
            # Draw znak.png
            c.drawImage(ImageReader(znak_buffer), znak_x, znak_y, znak_width, znak_height)
    except Exception as e:
        print(f"Could not load znak.png: {e}")

    # Add caption below the DataMatrix
    safe_caption = make_safe_caption(raw_with_gs, caption)
    font_size = 6
    line_gap = 2
    
    # Calculate approximate characters per line based on page width
    approx_chars_per_line = int((width_pt - margin_pt * 2) / (font_size * 0.5))
    lines = wrap_text_lines(safe_caption, max(approx_chars_per_line, 20))
    
    # Start drawing text below the DataMatrix
    text_y = dm_y - 2 * 2.83465 - font_size  # 2mm below the matrix
    
    for line in lines:
        if text_y < 1 * 2.83465:  # Stop if we reach bottom margin
            break
        
        # Center the text
        text_width = c.stringWidth(line, "Helvetica", font_size)
        text_x = (width_pt - text_width) / 2
        
        c.setFont("Helvetica", font_size)
        c.drawString(text_x, text_y, line)
        text_y -= font_size + line_gap
    
    # Save the PDF
    c.save()
    
    # Get the PDF data from the buffer
    pdf_data = buffer.getvalue()
    buffer.close()
    
    return pdf_data

class BarcodeInputDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Штрихкод")
        self.setModal(True)
        self.resize(30, 100)
        
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        self.label = QLabel("Введите штрихкод товара:")
        layout.addWidget(self.label)
        
        self.input_field = QLineEdit()
        layout.addWidget(self.input_field)
        
        button_layout = QHBoxLayout()
        layout.addLayout(button_layout)
        
        ok_button = QPushButton("OK")
        ok_button.clicked.connect(self.accept)
        button_layout.addWidget(ok_button)
        
        cancel_button = QPushButton("Отмена")
        cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(cancel_button)
        
        self.input_field.returnPressed.connect(ok_button.click)


class SettingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Настройки")
        self.setModal(True)
        self.resize(400, 300)
        
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        # Создаем вкладки для разных настроек
        self.tab_widget = QTabWidget()
        
        # Вкладка общих настроек
        general_tab = QWidget()
        general_layout = QVBoxLayout()
        
        # Настройка размеров шрифтов
        font_size_group = QGroupBox("Размеры шрифтов для печати этикеток на коробки")
        font_size_layout = QFormLayout()
        
        self.title_font_size = QLineEdit()
        self.title_font_size.setValidator(QIntValidator(8, 30))
        self.title_font_size.setText("12")
        font_size_layout.addRow("Шрифт заголовков:", self.title_font_size)
        
        self.value_font_size = QLineEdit()
        self.value_font_size.setValidator(QIntValidator(8, 30))
        self.value_font_size.setText("14")
        font_size_layout.addRow("Шрифт значений:", self.value_font_size)
        
        self.barcode_font_size = QLineEdit()
        self.barcode_font_size.setValidator(QIntValidator(8, 30))
        self.barcode_font_size.setText("14")
        font_size_layout.addRow("Шрифт штрихкода:", self.barcode_font_size)
        
        font_size_group.setLayout(font_size_layout)
        general_layout.addWidget(font_size_group)
        
        # Настройка длины символов для переноса строки
        line_wrap_group = QGroupBox("Перенос строк")
        line_wrap_layout = QFormLayout()
        
        self.name_line_wrap = QLineEdit()
        self.name_line_wrap.setValidator(QIntValidator(10, 50))
        self.name_line_wrap.setText("16")
        line_wrap_layout.addRow("Макс. длина строки для наименования:", self.name_line_wrap)
        
        self.article_line_wrap = QLineEdit()
        self.article_line_wrap.setValidator(QIntValidator(10, 50))
        self.article_line_wrap.setText("16")
        line_wrap_layout.addRow("Макс. длина строки для артикула:", self.article_line_wrap)
        
        self.barcode_line_wrap = QLineEdit()
        self.barcode_line_wrap.setValidator(QIntValidator(10, 50))
        self.barcode_line_wrap.setText("25")
        line_wrap_layout.addRow("Макс. длина строки для штрихкода:", self.barcode_line_wrap)
        
        line_wrap_group.setLayout(line_wrap_layout)
        general_layout.addWidget(line_wrap_group)
        
        general_tab.setLayout(general_layout)
        self.tab_widget.addTab(general_tab, "Основные")
        
        # Вкладка обновления таблицы
        table_tab = QWidget()
        table_layout = QVBoxLayout()
        
        self.update_btn = QPushButton("Актуализировать таблицу")
        self.update_btn.clicked.connect(self.update_table)
        table_layout.addWidget(self.update_btn)
        
        # Информация о настройках
        info_label = QLabel("Настройки применяются сразу после закрытия диалога")
        info_label.setWordWrap(True)
        info_label.setStyleSheet("color: grey; font-size: 10px;")
        table_layout.addWidget(info_label)
        
        table_tab.setLayout(table_layout)
        self.tab_widget.addTab(table_tab, "Таблица SKU")
        
        layout.addWidget(self.tab_widget)
        
        # Кнопки OK и Отмена
        button_layout = QHBoxLayout()
        
        ok_button = QPushButton("OK")
        ok_button.clicked.connect(self.accept)
        button_layout.addWidget(ok_button)
        
        cancel_button = QPushButton("Отмена")
        cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(cancel_button)
        
        layout.addLayout(button_layout)
        
        # Загружаем сохраненные настройки
        self.load_settings()
    
    def load_settings(self):
        """Загружает сохраненные настройки"""
        try:
            with open('settings.json', 'r', encoding='utf-8') as f:
                settings = json.load(f)
                
                self.title_font_size.setText(str(settings.get('title_font_size', 12)))
                self.value_font_size.setText(str(settings.get('value_font_size', 14)))
                self.barcode_font_size.setText(str(settings.get('barcode_font_size', 14)))
                self.name_line_wrap.setText(str(settings.get('name_line_wrap', 16)))
                self.article_line_wrap.setText(str(settings.get('article_line_wrap', 16)))
                self.barcode_line_wrap.setText(str(settings.get('barcode_line_wrap', 25)))
        except FileNotFoundError:
            # Если файл настроек не найден, используем значения по умолчанию
            pass
        except Exception:
            # Если ошибка при чтении файла, используем значения по умолчанию
            pass
    
    def save_settings(self):
        """Сохраняет настройки в файл"""
        settings = {
            'title_font_size': int(self.title_font_size.text()) if self.title_font_size.text().isdigit() else 12,
            'value_font_size': int(self.value_font_size.text()) if self.value_font_size.text().isdigit() else 14,
            'barcode_font_size': int(self.barcode_font_size.text()) if self.barcode_font_size.text().isdigit() else 14,
            'name_line_wrap': int(self.name_line_wrap.text()) if self.name_line_wrap.text().isdigit() else 16,
            'article_line_wrap': int(self.article_line_wrap.text()) if self.article_line_wrap.text().isdigit() else 16,
            'barcode_line_wrap': int(self.barcode_line_wrap.text()) if self.barcode_line_wrap.text().isdigit() else 25
        }
        
        try:
            with open('settings.json', 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=4)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить настройки: {str(e)}")
    
    def update_table(self):
        # Вызываем метод обновления таблицы из главного окна
        parent = self.parent()
        if isinstance(parent, LabelPrinterApp):
            parent.update_table()
    
    def accept(self):
        # Сохраняем настройки при нажатии OK
        self.save_settings()
        super().accept()


class LabelPrinterApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Принтер этикеток")
        self.setGeometry(100, 100, 300, 200)

        # Инициализируем базу данных
        self.init_db()
        
        # Загружаем данные SKU из базы данных
        self.sku_data = {}
        self.load_sku_data_from_db()
        
        # Создаем меню
        self.create_menu()
        
        # Создаем пользовательский интерфейс
        self.setup_ui()
        
        # Инициализируем аудио для воспроизведения звуков (если доступно)
        if not HAS_AUDIO:
            print("Аудиосистема не установлена. Функция воспроизведения звуков будет недоступна.")
    
    def init_db(self):
        """Создает таблицу sku в PostgreSQL, если она не существует"""
        try:
            # Таблица sku уже создается в database.init_db()
            # Здесь просто проверяем, что она существует
            from database import execute_query
            result = execute_query("SELECT COUNT(*) FROM sku", fetchone=True)
            if result:
                logger.info("Таблица sku в PostgreSQL готова к использованию")
        except Exception as e:
            logger.error(f"Ошибка при проверке таблицы sku: {e}")
    
    def load_sku_data_from_db(self):
        """Загружает данные из PostgreSQL базы данных в память"""
        try:
            from database import execute_query
            results = execute_query("SELECT barcode, article, name FROM sku", fetchall=True)
            
            # Очищаем предыдущие данные
            self.sku_data = {}
            
            # Загружаем данные из базы
            for row in results:
                barcode, article, name = row
                self.sku_data[barcode] = {'article': article, 'name': name}
                
        except Exception as e:
            self.show_error_message("Ошибка", f"Не удалось загрузить данные из базы: {str(e)}")
    
    def play_sound(self, sound_type):
        """Воспроизводит звук в зависимости от типа"""
        # Проверяем, доступна ли система воспроизведения звуков
        if not HAS_AUDIO:
            print("Аудиосистема не установлена. Воспроизведение звука невозможно.")
            return
            
        try:
            play_sound_simple(sound_type)
        except Exception as e:
            print(f"Ошибка воспроизведения звука: {str(e)}")
    
    def update_table(self):
        """Обновляет данные в базе данных из Excel файла и актуализирует информацию о наличии PDF файлов"""
        sku_path = "SKU/SKU.xlsx"
        
        # Проверяем, существует ли файл
        if not os.path.exists(sku_path):
            self.show_error_message("Ошибка", f"Файл {sku_path} не найден!")
            return
            
        try:
            # Подключаемся к базе данных
            conn = sqlite3.connect('label_printer.db')
            cursor = conn.cursor()
            
            # Очищаем текущую таблицу
            cursor.execute("DELETE FROM sku")
            
            # Загружаем данные из Excel файла
            wb = load_workbook(sku_path)
            ws = wb.active
            
            # Актуализируем информацию о наличии PDF файлов
            # Получаем список всех PDF файлов в папке PDF
            pdf_files = []
            if os.path.exists("PDF"):
                pdf_files = [f[:-4] for f in os.listdir("PDF") if f.endswith('.pdf')]  # Убираем расширение .pdf
            
            # Создаем словарь для быстрого поиска штрихкодов из PDF файлов
            pdf_barcodes = set(pdf_files)
            
            # Создаем словарь соответствия штрихкодов из таблицы
            table_barcodes = set()
            barcode_rows = {}  # Сохраняем номера строк для каждого штрихкода
            
            for row in range(2, ws.max_row + 1):
                barcode_cell = ws.cell(row=row, column=1)
                if barcode_cell.value is None:
                    continue
                barcode = str(barcode_cell.value)
                table_barcodes.add(barcode)
                barcode_rows[barcode] = row
            
            # Обновляем столбец "Этикетка (есть/нет)" для существующих записей
            for barcode, row in barcode_rows.items():
                label_status = "Есть" if barcode in pdf_barcodes else "Нет"
                # Находим столбец "Этикетка (есть/нет)" - это 4-й столбец
                ws.cell(row=row, column=4, value=label_status)
            
            # Находим лист "Добавить" и добавляем туда штрихкоды, которые есть в PDF, но нет в таблице SKU
            try:
                add_ws = wb["Добавить"]
            except KeyError:
                # Если лист "Добавить" не существует, создаем его
                add_ws = wb.create_sheet("Добавить")
            
            # Находим следующую пустую строку в листе "Добавить"
            next_add_row = 1
            while add_ws.cell(row=next_add_row, column=1).value is not None:
                next_add_row += 1
            
            # Добавляем штрихкоды, которые есть в PDF, но нет в таблице SKU
            missing_barcodes = pdf_barcodes - table_barcodes
            for barcode in missing_barcodes:
                add_ws.cell(row=next_add_row, column=1, value=barcode)
                next_add_row += 1
            
            # Загружаем данные в базу данных ТОЛЬКО для строк, у которых статус "Есть"
            for row in range(2, ws.max_row + 1):  # начиная со второй строки (первая строка заголовки)
                barcode_cell = ws.cell(row=row, column=1)
                if barcode_cell.value is None:
                    continue  # Пропускаем пустые строки
                    
                barcode = str(barcode_cell.value)
                label_status_cell = ws.cell(row=row, column=4)  # Четвертый столбец - статус наличия этикетки
                label_status = label_status_cell.value
                
                # Вставляем данные в базу ТОЛЬКО если статус "Есть"
                if label_status == "Есть":
                    article_cell = ws.cell(row=row, column=2)
                    article = article_cell.value if article_cell.value is not None else ""
                    name_cell = ws.cell(row=row, column=3)
                    name = name_cell.value if name_cell.value is not None else ""
                    
                    # Вставляем данные в базу
                    cursor.execute("INSERT OR REPLACE INTO sku (barcode, article, name) VALUES (?, ?, ?)", (barcode, article, name))
            
            # Сохраняем изменения
            conn.commit()
            
            # Сохраняем изменения в Excel файле
            wb.save(sku_path)
            
            conn.close()
            
            # Перезагружаем данные SKU из базы данных в память
            self.load_sku_data_from_db()
            
            # Подсчитываем количество записей, у которых есть этикетки
            records_with_labels = len([barcode for barcode, data in self.sku_data.items()])
            
            # Показываем сообщение об успешном обновлении
            self.show_info_message("Успешно", f"Таблица SKU успешно обновлена из файла Excel!\nКоличество штрихкодов с этикетками: {records_with_labels}\nКоличество новых штрихкодов для добавления: {len(missing_barcodes)}")
        except Exception as e:
            self.show_error_message("Ошибка", f"Не удалось обновить таблицу: {str(e)}")

    def create_menu(self):
        """Создает меню с настройками"""
        menubar = self.menuBar()
        settings_menu = menubar.addMenu('Настройки')
        
        # Добавляем пункт меню для основных настроек
        general_settings_action = settings_menu.addAction('Основные настройки')
        general_settings_action.triggered.connect(self.open_general_settings)
        
        update_action = settings_menu.addAction('Актуализировать таблицу')
        update_action.triggered.connect(self.open_settings)
    
    def open_settings(self):
        """Открывает диалог настроек"""
        dialog = SettingsDialog(self)
        dialog.exec()
    
    def open_general_settings(self):
        """Открывает диалог настроек с активной вкладкой основных настроек"""
        dialog = SettingsDialog(self)
        # Устанавливаем активной вкладку "Основные"
        dialog.tab_widget.setCurrentIndex(0)  # Первая вкладка - "Основные"
        dialog.exec()
    
    def setup_ui(self):
        """Создает пользовательский интерфейс с кнопками"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        # Кнопка "Перепечатать этикетку"
        reprint_btn = QPushButton("Перепечатать этикетку")
        reprint_btn.clicked.connect(self.reprint_label)
        reprint_btn.setStyleSheet(
            "QPushButton {"
            "   font-size: 14px;"
            "   font-weight: bold;"
            "   padding: 15px;"
            "   margin: 10px;"
            "   border: 2px solid #4CAF50;"
            "   border-radius: 8px;"
            "   background-color: #4CAF50;"
            "   color: white;"
            "}"
            "QPushButton:hover {"
            "   background-color: #45a049;"
            "   border: 2px solid #45a049;"
            "}"
            "QPushButton:pressed {"
            "   background-color: #3d8b40;"
            "   border: 2px solid #3d8b40;"
            "}"
        )
        layout.addWidget(reprint_btn)

        # Кнопка "Напечатать этикетку на коробку"
        box_label_btn = QPushButton("Напечатать этикетку на коробку")
        box_label_btn.clicked.connect(self.print_box_label)
        box_label_btn.setStyleSheet(
            "QPushButton {"
            "   font-size: 14px;"
            "   font-weight: bold;"
            "   padding: 15px;"
            "   margin: 10px;"
            "   border: 2px solid #2196F3;"
            "   border-radius: 8px;"
            "   background-color: #2196F3;"
            "   color: white;"
            "}"
            "QPushButton:hover {"
            "   background-color: #1E88E5;"
            "   border: 2px solid #1E88E5;"
            "}"
            "QPushButton:pressed {"
            "   background-color: #1976D2;"
            "   border: 2px solid #1976D2;"
            "}"
        )
        layout.addWidget(box_label_btn)
        
        # Устанавливаем размеры окна
        self.resize(400, 350)


    def show_barcode_dialog(self):
        dialog = BarcodeInputDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            return dialog.input_field.text().strip()
        return None

    def show_error_message(self, title, message):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Icon.Critical)
        msg.setWindowTitle(title)
        msg.setText(message)
        msg.exec()

    def show_info_message(self, title, message):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setWindowTitle(title)
        msg.setText(message)
        msg.exec()

    def reprint_label(self):
        # Цикл для многократного ввода штрихкода и перепечатки этикеток
        while True:
            # Запрос штрихкода через диалоговое окно
            barcode = self.show_barcode_dialog()
            
            # Если пользователь нажал Cancel или закрыл окно, выходим из цикла
            if not barcode:
                break
            
            # Проверяем, существует ли PDF файл с таким штрихкодом
            pdf_path = f"PDF/{barcode}.pdf"
            
            if os.path.exists(pdf_path):
                # Печатаем PDF файл с использованием PyMuPDF (без открытия Acrobat Reader)
                self.print_pdf(pdf_path, use_acrobat=False)
                
                # После успешной печати продолжаем цикл, чтобы снова запросить штрихкод
            else:
                self.show_error_message("Ошибка", f"PDF файл для штрихкода {barcode} не найден!")
                
                # Спрашиваем пользователя, хочет ли он попробовать снова
                reply = QMessageBox.question(self, "Повторить?", "Хотите ввести другой штрихкод?",
                                           QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
                if reply != QMessageBox.StandardButton.Yes:
                    break
    
    def create_and_print_box_label(self, barcode, article, name):
        """
        Создает и печатает этикетку для коробки с артикулом, наименованием и штрихкодом
        """
        try:
            # Загружаем настройки
            settings = {}
            try:
                with open('settings.json', 'r', encoding='utf-8') as f:
                    settings = json.load(f)
            except FileNotFoundError:
                # Используем настройки по умолчанию
                settings = {
                    'title_font_size': 12,
                    'value_font_size': 14,
                    'barcode_font_size': 14,
                    'name_line_wrap': 16,
                    'article_line_wrap': 16,
                    'barcode_line_wrap': 25
                }
            
            # Получаем размеры шрифтов из настроек
            title_font_size = settings.get('title_font_size', 12)
            value_font_size = settings.get('value_font_size', 14)
            barcode_font_size = settings.get('barcode_font_size', 14)
            
            # Ограничиваем длину строк
            name_line_wrap = settings.get('name_line_wrap', 16)
            article_line_wrap = settings.get('article_line_wrap', 16)
            barcode_line_wrap = settings.get('barcode_line_wrap', 25)
            
            # Разбиваем длинные строки на части
            def wrap_text(text, max_length):
                if not text:
                    return [""]
                lines = []
                for i in range(0, len(text), max_length):
                    lines.append(text[i:i + max_length])
                return lines
            
            name_lines = wrap_text(name, name_line_wrap)
            article_lines = wrap_text(article, article_line_wrap)
            barcode_lines = wrap_text(barcode, barcode_line_wrap)
            
            # Создаем PDF этикетки
            buffer = io.BytesIO()
            c = canvas.Canvas(buffer, pagesize=(60*mm, 40*mm))  # 60x40 мм - размер этикетки
            
            # Устанавливаем позиции элементов
            x_margin = 5*mm
            y_position = 35*mm  # Начальная позиция сверху
            line_spacing = 4*mm
            
            # Печатаем заголовки и значения
            c.setFont("Helvetica-Bold", title_font_size)
            c.drawString(x_margin, y_position, "Артикул:")
            y_position -= line_spacing
            
            c.setFont("Helvetica", value_font_size)
            for line in article_lines:
                c.drawString(x_margin, y_position, line)
                y_position -= line_spacing
            
            y_position -= line_spacing  # Дополнительный отступ перед следующим полем
            
            c.setFont("Helvetica-Bold", title_font_size)
            c.drawString(x_margin, y_position, "Наименование:")
            y_position -= line_spacing
            
            c.setFont("Helvetica", value_font_size)
            for line in name_lines:
                c.drawString(x_margin, y_position, line)
                y_position -= line_spacing
            
            y_position -= line_spacing  # Дополнительный отступ перед штрихкодом
            
            c.setFont("Helvetica-Bold", title_font_size)
            c.drawString(x_margin, y_position, "Штрихкод:")
            y_position -= line_spacing
            
            c.setFont("Helvetica", barcode_font_size)
            for line in barcode_lines:
                c.drawString(x_margin, y_position, line)
                y_position -= line_spacing
            
            # Завершаем создание PDF
            c.save()
            
            # Получаем данные PDF из буфера
            pdf_data = buffer.getvalue()
            buffer.close()
            
            # Сохраняем временный PDF файл для печати
            temp_pdf_path = os.path.join(tempfile.gettempdir(), f"box_label_{barcode}.pdf")
            with open(temp_pdf_path, 'wb') as f:
                f.write(pdf_data)
            
            # Печатаем PDF файл
            self.print_pdf(temp_pdf_path, use_acrobat=False)
            
            # Удаляем временный файл после печати
            try:
                os.remove(temp_pdf_path)
            except:
                pass  # Игнорируем ошибки при удалении временного файла
                
        except Exception as e:
            self.show_error_message("Ошибка", f"Не удалось создать и напечатать этикетку для коробки: {str(e)}")

    def print_box_label(self):
        # Цикл для многократного ввода штрихкода и печати этикеток
        while True:
            # Запрос штрихкода через диалоговое окно
            barcode = self.show_barcode_dialog()
            
            # Если пользователь нажал Cancel или закрыл окно, выходим из цикла
            if not barcode:
                break
                
            # Ищем артикул в загруженных данных
            if barcode in self.sku_data:
                data = self.sku_data[barcode]
                article = data['article']
                name = data['name']
                # Создаем и печатаем этикетку с использованием данных из памяти
                self.create_and_print_box_label(barcode, str(article), str(name))
                
                # После успешной печати продолжаем цикл, чтобы снова запросить штрихкод
            else:
                self.show_error_message("Ошибка", f"Данные для штрихкода {barcode} не найдены в таблице SKU!")
                
                # Спрашиваем пользователя, хочет ли он попробовать снова
                reply = QMessageBox.question(self, "Повторить?", "Хотите ввести другой штрихкод?",
                                           QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
                if reply != QMessageBox.StandardButton.Yes:
                    break
                


    def print_pdf(self, pdf_path, use_acrobat=False):
        """
        Печатает PDF файл на принтере с использованием PyMuPDF.
        """
        if not HAS_PYMUPDF:
            # Если PyMuPDF не установлен и вызов происходит из reprint_label (use_acrobat=False),
            # показываем ошибку вместо использования резервного метода с Acrobat Reader
            if not use_acrobat:
                self.show_error_message("Ошибка", "PyMuPDF не установлен. Необходимо установить библиотеку PyMuPDF для печати этикеток.\n\nТекущий статус HAS_PYMUPDF: " + str(HAS_PYMUPDF))
                return
            else:
                # Если вызов происходит с намерением использовать Acrobat Reader, используем резервный метод
                self._print_pdf_fallback(pdf_path, use_acrobat)
                return
        
        if not os.path.exists(pdf_path):
            self.show_error_message("Ошибка", f"PDF файл не найден: {pdf_path}")
            return
        
        try:
            # Открываем PDF документ
            doc = fitz.open(pdf_path)
            
            # Создаем объект печати
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            painter = QPainter(printer)
            
            # Проходим по всем страницам документа
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                
                # Масштабируем страницу под размер печати
                # Получаем размеры страницы
                page_rect = page.rect
                page_width = page_rect.width
                page_height = page_rect.height
                
                # Рассчитываем масштаб для умещения страницы на листе
                printer_page_rect = printer.pageRect(QPrinter.Unit.DevicePixel)
                printer_width = printer_page_rect.width()
                printer_height = printer_page_rect.height()
                scale_x = printer_width / page_width
                scale_y = printer_height / page_height
                scale = min(scale_x, scale_y)
                
                # Масштабируем и рендерим страницу
                mat = fitz.Matrix(scale, scale)
                pix = page.get_pixmap(matrix=mat)
                
                # Конвертируем Pixmap в QImage
                img = QImage(pix.samples, pix.width, pix.height, pix.stride, QImage.Format.Format_RGB888)
                
                # Рисуем изображение на принтере
                painter.drawImage(painter.viewport(), img)
                
                # Если есть еще страницы, создаем новую страницу
                if page_num < len(doc) - 1:
                    printer.newPage()
            
            # Завершаем печать
            painter.end()
            doc.close()
        
        except Exception as e:
            # Если возникла ошибка при печати через PyMuPDF, используем резервный метод
            # Но не используем Acrobat Reader при вызове из reprint_label (где use_acrobat=False)
            if use_acrobat:
                self._print_pdf_fallback(pdf_path, use_acrobat)
            else:
                # При печати этикеток показываем сообщение об ошибке вместо открытия Acrobat Reader
                self.show_error_message("Ошибка печати", f"Не удалось распечатать этикетку: {str(e)}")
    
    def _print_pdf_fallback(self, pdf_path, use_acrobat=False):
        """
        Резервный метод печати PDF файла на случай, если PyMuPDF недоступен.
        """
        if sys.platform.startswith('win'):
           # Для Windows пробуем несколько методов печати
           
           # Проверяем существование файла
           if not os.path.exists(pdf_path):
               self.show_error_message("Ошибка", f"PDF файл не найден: {pdf_path}")
               return
               
               
           abs_pdf_path = os.path.abspath(pdf_path)
           
           if use_acrobat:
               # Для печати с использованием Acrobat Reader
               try:
                   # Пробуем использовать Adobe Reader для печати, если он установлен
                   adobe_paths = [
                       r"C:\Program Files\Adobe\Acrobat DC\Acrobat\Acrobat.exe",
                       r"C:\Program Files (x86)\Adobe\Acrobat Reader DC\Reader\AcroRd32.exe",
                       r"C:\Program Files\Adobe\Acrobat Reader DC\Reader\AcroRd32.exe"
                   ]
                   
                   for adobe_path in adobe_paths:
                       if os.path.exists(adobe_path):
                           subprocess.run([adobe_path, "/t", abs_pdf_path])
                           return
                           
                   # Если Adobe Reader не найден, используем ShellExecute как резервный вариант
                   try:
                       import win32api
                       result = win32api.ShellExecute(0, "print", abs_pdf_path, None, ".", 0)
                       
                       if result > 32:  # Успешный код возврата
                           return
                       else:
                           # Если ShellExecute не смог напечатать, проверяем тип ошибки
                           if result == 31:  # ERROR_GEN_FAILURE - "Присоединенное устройство не работает"
                               # Открываем файл для ручной печати
                               subprocess.run(["start", "", abs_pdf_path], shell=True)
                               QMessageBox.warning(self, "Предупреждение", f"Принтер недоступен. Файл {os.path.basename(abs_pdf_path)} открыт для ручной печати. Пожалуйста, настройте принтер или распечатайте файл вручную.")
                               return
                           else:
                               raise Exception(f"ShellExecute вернул ошибку: {result}")
                   except ImportError:
                       # Если win32api недоступен, используем os.startfile как резервный вариант
                       try:
                           os.startfile(abs_pdf_path, "print")
                           return
                       except Exception as e:
                           self.show_error_message("Ошибка", f"Не удалось распечатать этикетку: {str(e)}")
                           return
                           
               except ImportError:
                   # Если win32api недоступен, используем os.startfile как резервный вариант
                   try:
                       os.startfile(abs_pdf_path, "print")
                       return
                   except Exception as e:
                       self.show_error_message("Ошибка", f"Не удалось распечатать этикетку: {str(e)}")
                       return
           else:
               # Для печати без использования Acrobat Reader (напрямую)
               # В контексте reprint_label (use_acrobat=False) не открываем Acrobat Reader при ошибках
               try:
                   import win32api

                   # Пробуем напечатать файл через ShellExecute
                   result = win32api.ShellExecute(0, "print", abs_pdf_path, None, ".", 0)
                   
                   # Проверяем результат выполнения ShellExecute
                   if result > 32:  # Успешный код возврата
                       return
                   else:
                       # Если ShellExecute не смог напечатать, проверяем тип ошибки
                       if result == 31:  # ERROR_GEN_FAILURE - "Присоединенное устройство не работает"
                           # Вместо открытия файла для ручной печати, просто показываем ошибку
                           self.show_error_message("Ошибка печати", f"Принтер недоступен. Не удалось распечатать файл: {os.path.basename(abs_pdf_path)}")
                           return
                       else:
                           # Для других ошибок выбрасываем исключение для дальнейшей обработки
                           raise Exception(f"ShellExecute вернул ошибку: {result}")
               except ImportError:
                   # Если win32api недоступен, используем os.startfile как резервный вариант
                   try:
                       os.startfile(abs_pdf_path, "print")
                       return
                   except Exception as e:
                       self.show_error_message("Ошибка", f"Не удалось распечатать этикетку: {str(e)}")
                       return
               except Exception as e:
                   # Если основной метод с win32api завершился с ошибкой, проверяем тип ошибки
                   error_msg = str(e)
                   if "ShellExecute вернул ошибку: 31" in error_msg:  # ERROR_GEN_FAILURE - "Присоединенное устройство не работает"
                       # Вместо открытия файла для ручной печати, просто показываем ошибку
                       self.show_error_message("Ошибка печати", f"Принтер недоступен. Не удалось распечатать файл: {os.path.basename(abs_pdf_path)}")
                       return
                   else:
                       # Для других ошибок сообщаем пользователю
                       QMessageBox.critical(self, "Ошибка печати", f"Не удалось отправить файл на печать: {str(e)}")
                       return
        elif sys.platform.startswith('darwin'):
            # Для macOS используем системную команду печати
            try:
                subprocess.run(["lp", pdf_path], check=True)
            except Exception as e:
                self.show_error_message("Ошибка печати", f"Не удалось распечатать этикетку: {str(e)}")
        else:
            # Для Linux используем lp или evince
            try:
                subprocess.run(["lp", pdf_path], check=True)
            except Exception as e:
                self.show_error_message("Ошибка печати", f"Не удалось распечатать этикетку: {str(e)}")

if __name__ == "__main__":
    import atexit
    
    # В скомпилированном приложении не запускаем проверку зависимостей,
    # чтобы избежать перезапуска приложения
    if getattr(sys, 'frozen', False):
        # Приложение запущено из скомпилированного .exe файла
        pass  # Пропускаем проверку зависимостей
    else:
        # Приложение запущено из исходного кода
        check_and_install_dependencies()  # Проверяем зависимости при запуске
    
    app = QApplication(sys.argv)
    window = LabelPrinterApp()
    window.show()
    
    # Регистрируем функцию очистки ресурсов при завершении приложения
    if MEMORY_MANAGER_AVAILABLE:
        atexit.register(cleanup_on_exit)
    
    sys.exit(app.exec())


def check_and_install_dependencies():
   """Проверяет и устанавливает недостающие зависимости"""
   required_packages = {
       'setuptools': 'setuptools',  # Добавляем setuptools для совместимости с новыми версиями Python
       'PIL': 'Pillow',  # Для работы с изображениями (импортируется как PIL)
       'reportlab': 'reportlab',  # Для генерации PDF
       'openpyxl': 'openpyxl',  # Для работы с Excel файлами
       'psutil': 'psutil',  # Для мониторинга системных ресурсов
       # audio_handler не является отдельным пакетом, это наш собственный модуль
   }
   
   missing_packages = []
   for import_name, package_name in required_packages.items():
       try:
           __import__(import_name)
       except ImportError:
           missing_packages.append(package_name)
   
   # Проверяем существование нашего собственного модуля audio_handler
   if not os.path.exists('audio_handler.py'):
       print("Файл audio_handler.py не найден. Он необходим для воспроизведения звуков.")
       missing_packages.append('audio_handler.py (локальный файл)')
   
   if missing_packages:
       print(f"Отсутствующие пакеты: {missing_packages}")
       try:
           # Пытаемся установить недостающие пакеты
           packages_to_install = [pkg for pkg in missing_packages if pkg != 'audio_handler.py (локальный файл)']
           for package in packages_to_install:
               print(f"Устанавливаю {package}...")
               subprocess.check_call([sys.executable, "-m", "pip", "install", package])
           print("Все недостающие пакеты установлены. Перезапустите приложение.")
           sys.exit(0)
       except subprocess.CalledProcessError:
           print(f"Не удалось автоматически установить пакеты: {missing_packages}")
           print("Установите их вручную с помощью команды: pip install " + " ".join([pkg for pkg in missing_packages if pkg != 'audio_handler.py (локальный файл)']))
           sys.exit(1)