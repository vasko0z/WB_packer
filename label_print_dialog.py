import sys
import os
import json
import sqlite3
import tempfile
from pathlib import Path

from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                            QPushButton, QLabel, QLineEdit, QDialog, QMessageBox, QMenu, QMenuBar,
                            QDialogButtonBox, QTabWidget, QGroupBox, QFormLayout, QSpinBox,
                            QFileDialog, QProgressDialog, QInputDialog, QPlainTextEdit, QGridLayout)
from PyQt6.QtCore import Qt, QSizeF, QThread, pyqtSignal
from PyQt6.QtPrintSupport import QPrinter, QPrintDialog
from PyQt6.QtGui import QFont, QPageSize, QPageLayout, QPainter, QPen, QImage, QPixmap, QIntValidator
import pandas as pd

# Import PyMuPDF (fitz) for PDF handling - optional dependency with fallback
try:
    import fitz  # type: ignore # PyMuPDF
    HAS_PYMUPDF = True
except ImportError as e:
    HAS_PYMUPDF = False

# Import reportlab for PDF generation
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    HAS_REPORTLAB = True
except ImportError as e:
    HAS_REPORTLAB = False
    canvas = None
    mm = None
    pdfmetrics = None
    TTFont = None

def print_pdf(pdf_path, use_acrobat=False):
    """
    Печатает PDF файл на принтере с использованием PyMuPDF или резервных методов.
    """
    import logging
    logger = logging.getLogger(__name__)
    
    # Проверяем доступность PyMuPDF
    if not HAS_PYMUPDF:
        logger.warning("PyMuPDF не доступен (HAS_PYMUPDF=False), используем резервный метод")
        return _print_pdf_fallback(pdf_path, use_acrobat)
    
    # Дополнительная проверка импорта
    try:
        import fitz
        logger.debug(f"PyMuPDF версия: {fitz.version}")
    except Exception as e:
        logger.warning(f"Не удалось импортировать fitz: {e}, используем резервный метод")
        return _print_pdf_fallback(pdf_path, use_acrobat)

    if not os.path.exists(pdf_path):
        QMessageBox.critical(None, "Ошибка", f"PDF файл не найден: {pdf_path}")
        return False

    try:
        # Открываем PDF документ
        doc = fitz.open(pdf_path)

        # Создаем объект печати
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        painter = QPainter(printer)

        # Проходим по всем страницам документа
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)

            # Масштабируем страницу под размер печати
            # Получаем размеры страницы
            page_rect = page.rect
            page_width = page_rect.width
            page_height = page_rect.height

            # Рассчитываем масштаб для умещения страницы на листе
            printer_page_rect = printer.pageRect(QPrinter.Unit.DevicePixel)
            printer_width = printer_page_rect.width()
            printer_height = printer_page_rect.height()
            scale_x = printer_width / page_width
            scale_y = printer_height / page_height
            scale = min(scale_x, scale_y)

            # Масштабируем и рендерим страницу
            mat = fitz.Matrix(scale, scale)
            pix = page.get_pixmap(matrix=mat)

            # Конвертируем Pixmap в QImage
            img = QImage(pix.samples, pix.width, pix.height, pix.stride, QImage.Format.Format_RGB888)

            # Рисуем изображение на принтере
            painter.drawImage(painter.viewport(), img)

            # Если есть еще страницы, создаем новую страницу
            if page_num < len(doc) - 1:
                printer.newPage()

        # Завершаем печать
        painter.end()
        doc.close()
        return True

    except Exception as e:
        # Если возникла ошибка при печати через PyMuPDF, используем резервный метод
        logger.error(f"Ошибка печати через PyMuPDF: {e}, используем резервный метод")
        return _print_pdf_fallback(pdf_path, use_acrobat)


def _print_pdf_fallback(pdf_path, use_acrobat=False):
    """
    Резервный метод печати PDF файла на случай, если PyMuPDF недоступен.
    """
    # Проверяем существование файла
    if not os.path.exists(pdf_path):
        QMessageBox.critical(None, "Ошибка", f"PDF файл не найден: {pdf_path}")
        return False
        
    abs_pdf_path = os.path.abspath(pdf_path)
    
    if sys.platform.startswith('win'):
       # Для Windows пробуем несколько методов печати
       
       if use_acrobat:
           # Для печати с использованием Acrobat Reader
           try:
               # Пробуем использовать Adobe Reader для печати, если он установлен
               adobe_paths = [
                   r"C:\Program Files\Adobe\Acrobat DC\Acrobat\Acrobat.exe",
                   r"C:\Program Files (x86)\Adobe\Acrobat Reader DC\Reader\AcroRd32.exe",
                   r"C:\Program Files\Adobe\Acrobat Reader DC\Reader\AcroRd32.exe"
               ]
               
               for adobe_path in adobe_paths:
                   if os.path.exists(adobe_path):
                       import subprocess
                       subprocess.run([adobe_path, "/t", abs_pdf_path])
                       return True
                       
               # Если Adobe Reader не найден, используем ShellExecute как резервный вариант
               try:
                   import win32api
                   result = win32api.ShellExecute(0, "print", abs_pdf_path, None, ".", 0)
                   
                   if result > 32:  # Успешный код возврата
                       return True
                   else:
                       # Если ShellExecute не смог напечатать, проверяем тип ошибки
                       if result == 31:  # ERROR_GEN_FAILURE - "Присоединенное устройство не работает"
                           # Открываем файл для ручной печати
                           subprocess.run(["start", "", abs_pdf_path], shell=True)
                           QMessageBox.warning(None, "Предупреждение", f"Принтер недоступен. Файл {os.path.basename(abs_pdf_path)} открыт для ручной печати. Пожалуйста, настройте принтер или распечатайте файл вручную.")
                           return True
                       else:
                           raise Exception(f"ShellExecute вернул ошибку: {result}")
               except ImportError:
                   # Если win32api недоступен, используем os.startfile как резервный вариант
                   try:
                       os.startfile(abs_pdf_path, "print")
                       return True
                   except OSError as e:
                       # Если не удается напечатать, пробуем открыть файл
                       if "not associated" in str(e).lower() or "сопоставлено" in str(e).lower():
                           try:
                               os.startfile(abs_pdf_path)  # Открываем файл вместо печати
                               QMessageBox.warning(None, "Предупреждение", f"Файл {os.path.basename(abs_pdf_path)} открыт для просмотра. Пожалуйста, распечатайте его вручную.")
                               return True
                           except Exception:
                               pass
                       QMessageBox.critical(None, "Ошибка", f"Не удалось распечатать этикетку: {str(e)}\n\nПопробуйте установить PDF-ридер (например, Adobe Reader) как приложение по умолчанию для PDF файлов.")
                       return False
                       
           except ImportError:
               # Если win32api недоступен, используем os.startfile как резервный вариант
               try:
                   os.startfile(abs_pdf_path, "print")
                   return True
               except Exception as e:
                   QMessageBox.critical(None, "Ошибка", f"Не удалось распечатать этикетку: {str(e)}")
                   return False
       else:
           # Для печати без использования Acrobat Reader (напрямую)
           # В контексте reprint_label (use_acrobat=False) не открываем Acrobat Reader при ошибках
           try:
               import win32api

               # Пробуем напечатать файл через ShellExecute
               result = win32api.ShellExecute(0, "print", abs_pdf_path, None, ".", 0)
               
               # Проверяем результат выполнения ShellExecute
               if result > 32:  # Успешный код возврата
                   return True
               else:
                   # Если ShellExecute не смог напечатать, проверяем тип ошибки
                   if result == 31:  # ERROR_GEN_FAILURE - "Присоединенное устройство не работает"
                       # Вместо открытия файла для ручной печати, просто показываем ошибку
                       QMessageBox.critical(None, "Ошибка печати", f"Принтер недоступен. Не удалось распечатать файл: {os.path.basename(abs_pdf_path)}")
                       return False
                   else:
                       # Для других ошибок выбрасываем исключение для дальнейшей обработки
                       raise Exception(f"ShellExecute вернул ошибку: {result}")
           except ImportError:
               # Если win32api недоступен, используем os.startfile как резервный вариант
               try:
                   os.startfile(abs_pdf_path, "print")
                   return True
               except OSError as e:
                   # Если не удается напечатать, пробуем открыть файл
                   if "not associated" in str(e).lower() or "сопоставлено" in str(e).lower():
                       try:
                           os.startfile(abs_pdf_path)  # Открываем файл вместо печати
                           QMessageBox.warning(None, "Предупреждение", f"Файл {os.path.basename(abs_pdf_path)} открыт для просмотра. Пожалуйста, распечатайте его вручную.")
                           return True
                       except Exception:
                           pass
                   QMessageBox.critical(None, "Ошибка", f"Не удалось распечатать этикетку: {str(e)}\n\nПопробуйте установить PDF-ридер (например, Adobe Reader) как приложение по умолчанию для PDF файлов.")
                   return False
           except Exception as e:
               # Если основной метод с win32api завершился с ошибкой, проверяем тип ошибки
               error_msg = str(e)
               if "ShellExecute вернул ошибку: 31" in error_msg:  # ERROR_GEN_FAILURE - "Присоединенное устройство не работает"
                   # Вместо открытия файла для ручной печати, просто показываем ошибку
                   QMessageBox.critical(None, "Ошибка печати", f"Принтер недоступен. Не удалось распечатать файл: {os.path.basename(abs_pdf_path)}")
                   return False
               else:
                   # Для других ошибок сообщаем пользователю
                   QMessageBox.critical(None, "Ошибка печати", f"Не удалось отправить файл на печать: {str(e)}")
                   return False
    elif sys.platform.startswith('darwin'):
        # Для macOS используем системную команду печати
        try:
            import subprocess
            subprocess.run(["lp", pdf_path], check=True)
            return True
        except Exception as e:
            QMessageBox.critical(None, "Ошибка печати", f"Не удалось распечатать этикетку: {str(e)}")
            return False
    else:
        # Для Linux используем lp или evince
        try:
            import subprocess
            subprocess.run(["lp", pdf_path], check=True)
            return True
        except Exception as e:
            QMessageBox.critical(None, "Ошибка печати", f"Не удалось распечатать этикетку: {str(e)}")
            return False


def print_existing_label(parent, barcode):
    """
    Печатает существующую этикетку из PDF файла
    """
    try:
        # Проверяем, существует ли PDF файл с таким штрихкодом
        pdf_path = f"PDF/{barcode}.pdf"
        
        if os.path.exists(pdf_path):
            return print_pdf(pdf_path, use_acrobat=False)
        else:
            return False
    except Exception as e:
        QMessageBox.critical(parent, "Ошибка", f"Ошибка при печати этикетки: {str(e)}")
        return False


def create_and_print_transport_label(printer, destination, box_id):
   """
   Создает и печатает транспортную этикетку для коробки с новым форматом:
   
   НАЗВАНИЕ
   
   Коробка
   1
   """
   try:
       import io
       import os
       import tempfile
       from reportlab.pdfgen import canvas
       from reportlab.lib.units import mm
       from reportlab.pdfbase import pdfmetrics
       from reportlab.pdfbase.ttfonts import TTFont
       
       # Попробуем использовать шрифт, который поддерживает кириллицу
       font_name_regular = 'Helvetica'
       font_name_bold = 'Helvetica-Bold'
       
       try:
           font_path = None
           possible_fonts = [
               "arial.ttf", "Arial.ttf", "ARIAL.TTF",
               "LiberationSans-Regular.ttf", "DejaVuSans.ttf",
               "calibri.ttf", "Calibri.ttf", "CALIBRI.TTF"
           ]
           
           for font_name in possible_fonts:
               if os.path.exists(font_name):
                   font_path = font_name
                   break
               system_font_paths = [
                   r"C:\Windows\Fonts\\" + font_name,
                   r"/usr/share/fonts/truetype/dejavu/" + font_name,
                   r"/System/Library/Fonts/" + font_name
               ]
               for path in system_font_paths:
                   if os.path.exists(path):
                       font_path = path
                       break
               if font_path:
                   break
           
           if font_path:
               pdfmetrics.registerFont(TTFont('CyrillicRegular', font_path))
               font_name_regular = 'CyrillicRegular'
               
               bold_font_path = None
               possible_bold_fonts = [
                   font_path.replace('.ttf', ' Bold.ttf').replace('.TTF', ' Bold.TTF'),
                   font_path.replace('Regular', 'Bold'),
                   font_path.replace('regular', 'bold'),
                   "arialbd.ttf", "Arial Bold.ttf", "ARIALBD.TTF",
                   "LiberationSans-Bold.ttf", "DejaVuSans-Bold.ttf"
               ]
               
               for bold_path in possible_bold_fonts:
                   if os.path.exists(bold_path):
                       bold_font_path = bold_path
                       break
                   for path in system_font_paths:
                       if 'Fonts' in path:
                           bold_sys_path = path.replace(os.path.basename(path), os.path.basename(bold_path))
                           if os.path.exists(bold_sys_path):
                               bold_font_path = bold_sys_path
                               break
                   if bold_font_path:
                       break
               
               if bold_font_path:
                   pdfmetrics.registerFont(TTFont('CyrillicBold', bold_font_path))
                   font_name_bold = 'CyrillicBold'
               else:
                   font_name_bold = 'CyrillicRegular'
       except:
           pass
       
       buffer = io.BytesIO()
       # Стандартный размер этикетки 58x40 мм
       c = canvas.Canvas(buffer, pagesize=(58*mm, 40*mm))
       
       # Центрируем весь контент по вертикали
       # Общая высота содержимого: название (2 строки) + отступ + "Коробка" + отступ + номер
       # Примерно: 12pt + 12pt + 5mm + 12pt + 5mm + 24pt = ~60mm
       # Центральная позиция для вертикального центрирования
       
       # Рисуем название направления (Поставка) - крупный шрифт, центрировано
       c.setFont(font_name_bold, 18)
       dest_text = str(destination)
       
       # Если название слишком длинное, уменьшаем шрифт
       if len(dest_text) > 14:
           c.setFont(font_name_bold, 16)
       if len(dest_text) > 16:
           c.setFont(font_name_bold, 14)
       if len(dest_text) > 20:
           c.setFont(font_name_bold, 12)
           
       c.drawCentredString(29*mm, 28*mm, dest_text)
       
       # Отступ
       
       # Рисуем "Коробка" - средний шрифт
       c.setFont(font_name_regular, 14)
       c.drawCentredString(29*mm, 18*mm, "Коробка")
       
       # Рисуем номер коробки - крупный шрифт
       c.setFont(font_name_bold, 24)
       # Извлекаем только номер из "Коробка-1" или оставляем как есть
       box_num = str(box_id)
       if "-" in box_num:
           box_num = box_num.split("-")[-1]
       elif " " in box_num:
           box_num = box_num.split(" ")[-1]
       
       c.drawCentredString(29*mm, 8*mm, box_num)
       
       c.save()
       
       pdf_data = buffer.getvalue()
       buffer.close()
       
       temp_pdf_path = os.path.join(tempfile.gettempdir(), f"transport_label_{box_id}.pdf")
       with open(temp_pdf_path, 'wb') as f:
           f.write(pdf_data)
       
       result = print_pdf(temp_pdf_path, use_acrobat=False)
       
       try:
           os.remove(temp_pdf_path)
       except:
           pass
           
       return result
       
   except Exception as e:
       from PyQt6.QtWidgets import QMessageBox
       QMessageBox.critical(None, "Ошибка", f"Ошибка при создании транспортной этикетки: {str(e)}")
       return False


def create_and_print_box_label(printer, barcode, article, name):
   """
   Создает и печатает этикетку для коробки с заданными параметрами
   """
   try:
       # Загружаем настройки из общей базы данных
       settings = {}
       try:
           from database import execute_query
           result = execute_query(
               "SELECT value FROM app_settings WHERE key = 'label_print_settings'",
               fetchone=True
           )
           
           if result:
               import json
               settings = json.loads(result[0])
           else:
               # Если настройки не найдены в базе, используем настройки по умолчанию
               settings = {
                   'title_font_size': 12,
                   'value_font_size': 14,
                   'barcode_font_size': 14,
                   'name_line_wrap': 16,
                   'article_line_wrap': 16
               }
       except Exception as e:
           # Если возникла ошибка при загрузке из базы, используем настройки по умолчанию
           settings = {
               'title_font_size': 12,
               'value_font_size': 14,
               'barcode_font_size': 14,
               'name_line_wrap': 16,
               'article_line_wrap': 16
           }
       
       # Получаем размеры шрифтов из настроек
       title_font_size = settings.get('title_font_size', 12)
       value_font_size = settings.get('value_font_size', 14)
       barcode_font_size = settings.get('barcode_font_size', 14)
       
       # Ограничиваем длину строк
       name_line_wrap = settings.get('name_line_wrap', 16)
       article_line_wrap = settings.get('article_line_wrap', 16)
       # Убираем использование настройки barcode_line_wrap, так как она больше не используется
       
       # Разбиваем длинные строки на части
       def wrap_text(text, max_length):
           if not text:
               return [""]
           lines = []
           for i in range(0, len(text), max_length):
               lines.append(text[i:i + max_length])
           return lines
       
       name_lines = wrap_text(name, name_line_wrap)
       article_lines = wrap_text(article, article_line_wrap)
       # Для штрихкода не применяем перенос строк, выводим как есть
       barcode_lines = [barcode]
       
       # Создаем PDF этикетки
       import io
       
       # Проверяем, доступен ли reportlab
       if not HAS_REPORTLAB:
           QMessageBox.critical(None, "Ошибка", "Модуль reportlab не установлен. Установите его с помощью: pip install reportlab")
           return False
       
       # Попробуем использовать шрифт, который поддерживает кириллицу
       try:
           # Попробуем использовать шрифт Arial Unicode MS или Liberation Sans
           font_path = None
           possible_fonts = [
               "arial.ttf", "Arial.ttf", "ARIAL.TTF",
               "LiberationSans-Regular.ttf", "DejaVuSans.ttf",
               "calibri.ttf", "Calibri.ttf", "CALIBRI.TTF"
           ]
           
           for font_name in possible_fonts:
               if os.path.exists(font_name):
                   font_path = font_name
                   break
               # Также проверим в системных каталогах
               system_font_paths = [
                   r"C:\Windows\Fonts\%s" % font_name,
                   r"/usr/share/fonts/truetype/dejavu/%s" % font_name,
                   r"/System/Library/Fonts/%s" % font_name
               ]
               for path in system_font_paths:
                   if os.path.exists(path):
                       font_path = path
                       break
               if font_path:
                   break
           
           if font_path:
               # Регистрируем шрифты
               pdfmetrics.registerFont(TTFont('CyrillicRegular', font_path))
               
               # Попробуем найти шрифт для жирного начертания
               bold_font_path = None
               possible_bold_fonts = [
                   font_path.replace('.ttf', ' Bold.ttf').replace('.TTF', ' Bold.TTF'),
                   font_path.replace('Regular', 'Bold'),
                   font_path.replace('regular', 'bold'),
                   "arialbd.ttf", "Arial Bold.ttf", "ARIALBD.TTF",
                   "LiberationSans-Bold.ttf", "DejaVuSans-Bold.ttf"
               ]
               
               for bold_path in possible_bold_fonts:
                   if os.path.exists(bold_path):
                       bold_font_path = bold_path
                       break
                   # Проверим системные каталоги для жирного шрифта
                   for path in system_font_paths:
                       if 'Fonts' in path:
                           bold_sys_path = path.replace(os.path.basename(path), os.path.basename(bold_path))
                           if os.path.exists(bold_sys_path):
                               bold_font_path = bold_sys_path
                               break
                   if bold_font_path:
                       break
               
               if bold_font_path:
                   pdfmetrics.registerFont(TTFont('CyrillicBold', bold_font_path))
                   font_name_regular = 'CyrillicRegular'
                   font_name_bold = 'CyrillicBold'
               else:
                   # Если жирный шрифт не найден, используем обычный для обоих случаев
                   font_name_regular = 'CyrillicRegular'
                   font_name_bold = 'CyrillicRegular'  # Используем обычный шрифт, но позже вручную сделаем жирным
           else:
               # Если не найден подходящий шрифт, используем стандартные, но с дополнительной обработкой
               font_name_regular = 'Helvetica'
               font_name_bold = 'Helvetica-Bold'
       except:
           # В случае ошибки при регистрации шрифта, используем стандартные шрифты
           font_name_regular = 'Helvetica'
           font_name_bold = 'Helvetica-Bold'
       
       buffer = io.BytesIO()
       c = canvas.Canvas(buffer, pagesize=(60*mm, 40*mm))  # 60x40 мм - размер этикетки
       
       # Устанавливаем позиции элементов
       x_margin = 5*mm
       y_position = 35*mm  # Начальная позиция сверху
       line_spacing = 4*mm
       
       # Печатаем заголовки и значения с поддержкой кириллицы
       c.setFont(font_name_bold, title_font_size)
       # Используем encode/decode для правильной обработки кириллических символов
       header_article = "Артикул:".encode('utf-8').decode('utf-8')
       c.drawString(x_margin, y_position, header_article)
       y_position -= line_spacing
       
       c.setFont(font_name_regular, value_font_size)
       for line in article_lines:
           # Убедимся, что строка в правильной кодировке
           safe_line = line.encode('utf-8').decode('utf-8') if isinstance(line, str) else str(line)
           c.drawString(x_margin, y_position, safe_line)
           y_position -= line_spacing
       
       y_position -= line_spacing  # Дополнительный отступ перед следующим полем
       
       c.setFont(font_name_bold, title_font_size)
       header_name = "Наименование:".encode('utf-8').decode('utf-8')
       c.drawString(x_margin, y_position, header_name)
       y_position -= line_spacing
       
       c.setFont(font_name_regular, value_font_size)
       for line in name_lines:
           # Убедимся, что строка в правильной кодировке
           safe_line = line.encode('utf-8').decode('utf-8') if isinstance(line, str) else str(line)
           c.drawString(x_margin, y_position, safe_line)
           y_position -= line_spacing
       
       y_position -= line_spacing  # Дополнительный отступ перед штрихкодом
       
       # Убрали надпись "Штрихкод:" согласно требованиям
       # Выводим только цифровой штрихкод без надписи "Штрихкод:"
       c.setFont(font_name_regular, barcode_font_size)
       for line in barcode_lines:
           # Выводим только цифровой штрихкод
           safe_line = line.encode('utf-8').decode('utf-8') if isinstance(line, str) else str(line)
           c.drawString(x_margin, y_position, safe_line)
           y_position -= line_spacing
       
       # Завершаем создание PDF
       c.save()
       
       # Получаем данные PDF из буфера
       pdf_data = buffer.getvalue()
       buffer.close()
       
       # Сохраняем временный PDF файл для печати
       temp_pdf_path = os.path.join(tempfile.gettempdir(), f"box_label_{barcode}.pdf")
       with open(temp_pdf_path, 'wb') as f:
           f.write(pdf_data)
       
       # Печатаем PDF файл
       result = print_pdf(temp_pdf_path, use_acrobat=False)
       
       # Удаляем временный файл после печати
       try:
           os.remove(temp_pdf_path)
       except:
           pass  # Игнорируем ошибки при удалении временного файла
       
       return result
       
   except Exception as e:
       QMessageBox.critical(None, "Ошибка", f"Ошибка при создании этикетки: {str(e)}")
       return False


class TextLabelDialog(QDialog):
    """Диалог для ввода текста и размера шрифта для печати этикетки"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Печать текста на этикетке")
        self.setModal(True)
        self.resize(400, 250)

        layout = QVBoxLayout()
        self.setLayout(layout)

        # Поле ввода текста (многострочное)
        text_layout = QVBoxLayout()
        self.text_input = QPlainTextEdit()
        self.text_input.setPlaceholderText("Введите текст для этикетки\n(поддерживается несколько строк)")
        self.text_input.setMaximumHeight(100)
        self.text_input.setStyleSheet(
            "QPlainTextEdit {"
            "   font-size: 14px;"
            "   padding: 10px;"
            "   border: 2px solid #ccc;"
            "   border-radius: 4px;"
            "}"
        )
        text_layout.addWidget(QLabel("Текст:"))
        text_layout.addWidget(self.text_input)
        layout.addLayout(text_layout)

        # Выбор размера шрифта
        font_layout = QHBoxLayout()
        self.font_size_spin = QSpinBox()
        self.font_size_spin.setRange(8, 72)
        self.font_size_spin.setValue(24)
        self.font_size_spin.setStyleSheet(
            "QSpinBox {"
            "   font-size: 14px;"
            "   padding: 5px;"
            "   border: 2px solid #ccc;"
            "   border-radius: 4px;"
            "}"
        )
        font_layout.addWidget(QLabel("Размер шрифта:"))
        font_layout.addWidget(self.font_size_spin)
        font_layout.addWidget(QLabel("пт"))
        font_layout.addStretch()
        layout.addLayout(font_layout)

        # Кнопки
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        self.print_btn = QPushButton("Печать")
        self.print_btn.clicked.connect(self.accept)
        self.print_btn.setStyleSheet(
            "QPushButton {"
            "   font-size: 14px;"
            "   font-weight: bold;"
            "   padding: 10px 20px;"
            "   border: 2px solid #4CAF50;"
            "   border-radius: 4px;"
            "   background-color: #4CAF50;"
            "   color: white;"
            "}"
            "QPushButton:hover {"
            "   background-color: #45a049;"
            "}"
        )
        btn_layout.addWidget(self.print_btn)

        self.cancel_btn = QPushButton("Отмена")
        self.cancel_btn.clicked.connect(self.reject)
        self.cancel_btn.setStyleSheet(
            "QPushButton {"
            "   font-size: 14px;"
            "   padding: 10px 20px;"
            "   border: 2px solid #757575;"
            "   border-radius: 4px;"
            "   background-color: #757575;"
            "   color: white;"
            "}"
            "QPushButton:hover {"
            "   background-color: #616161;"
            "}"
        )
        btn_layout.addWidget(self.cancel_btn)

        layout.addLayout(btn_layout)

    def get_text(self):
        """Получить введённый текст"""
        return self.text_input.toPlainText().strip()

    def get_font_size(self):
        """Получить размер шрифта"""
        return self.font_size_spin.value()


class BarcodeQuantityDialog(QDialog):
    """Диалог для ввода штрихкода и количества этикеток"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Печать этикетки на товар")
        self.setModal(True)
        self.setFixedSize(350, 180)
        
        layout = QGridLayout()
        self.setLayout(layout)
        
        # Поле для ввода штрихкода
        layout.addWidget(QLabel("Штрихкод:"), 0, 0)
        self.barcode_edit = QLineEdit()
        self.barcode_edit.setPlaceholderText("Введите штрихкод товара")
        layout.addWidget(self.barcode_edit, 0, 1)
        
        # Поле для ввода количества
        layout.addWidget(QLabel("Количество:"), 1, 0)
        self.quantity_spin = QSpinBox()
        self.quantity_spin.setRange(1, 999)
        self.quantity_spin.setValue(1)
        layout.addWidget(self.quantity_spin, 1, 1)
        
        # Кнопки OK и Отмена
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box, 2, 0, 1, 2)
        
        # Фокус на поле штрихкода
        self.barcode_edit.setFocus()
    
    def get_barcode(self):
        return self.barcode_edit.text().strip()
    
    def get_quantity(self):
        return self.quantity_spin.value()


class LabelPrintDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Печать этикеток")
        self.setModal(True)
        self.resize(400, 250)

        layout = QVBoxLayout()
        self.setLayout(layout)

        # Кнопка "Печать этикетки на товар"
        self.product_label_btn = QPushButton("Печать этикетки на товар")
        self.product_label_btn.clicked.connect(self.print_product_label)
        self.product_label_btn.setStyleSheet(
            "QPushButton {"
            "   font-size: 14px;"
            "   font-weight: bold;"
            "   padding: 15px;"
            "   margin: 10px;"
            "   border: 2px solid #4CAF50;"
            "   border-radius: 8px;"
            "   background-color: #4CAF50;"
            "   color: white;"
            "}"
            "QPushButton:hover {"
            "   background-color: #45a049;"
            "   border: 2px solid #45a049;"
            "}"
            "QPushButton:pressed {"
            "   background-color: #3d8b40;"
            "   border: 2px solid #3d8b40;"
            "}"
        )
        layout.addWidget(self.product_label_btn)

        # Кнопка "Печать этикетки на коробку"
        self.box_label_btn = QPushButton("Печать этикетки на коробку")
        self.box_label_btn.clicked.connect(self.print_box_label)
        self.box_label_btn.setStyleSheet(
            "QPushButton {"
            "   font-size: 14px;"
            "   font-weight: bold;"
            "   padding: 15px;"
            "   margin: 10px;"
            "   border: 2px solid #2196F3;"
            "   border-radius: 8px;"
            "   background-color: #2196F3;"
            "   color: white;"
            "}"
            "QPushButton:hover {"
            "   background-color: #1E88E5;"
            "   border: 2px solid #1E88E5;"
            "}"
            "QPushButton:pressed {"
            "   background-color: #1976D2;"
            "   border: 2px solid #1976D2;"
            "}"
        )
        layout.addWidget(self.box_label_btn)

        # Кнопка "Печать текста"
        self.text_label_btn = QPushButton("Печать текста")
        self.text_label_btn.clicked.connect(self.print_text_label)
        self.text_label_btn.setStyleSheet(
            "QPushButton {"
            "   font-size: 14px;"
            "   font-weight: bold;"
            "   padding: 15px;"
            "   margin: 10px;"
            "   border: 2px solid #FF9800;"
            "   border-radius: 8px;"
            "   background-color: #FF9800;"
            "   color: white;"
            "}"
            "QPushButton:hover {"
            "   background-color: #F57C00;"
            "   border: 2px solid #F57C00;"
            "}"
            "QPushButton:pressed {"
            "   background-color: #EF6C00;"
            "   border: 2px solid #EF6C00;"
            "}"
        )
        layout.addWidget(self.text_label_btn)

    def print_text_label(self):
        """Печать этикетки с произвольным текстом"""
        try:
            # Создаем диалог для ввода текста
            dialog = TextLabelDialog(self)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                text = dialog.get_text()
                font_size = dialog.get_font_size()
                if text:
                    self.print_custom_text(text, font_size)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Ошибка при печати текстовой этикетки: {e}", exc_info=True)
            QMessageBox.critical(self, "Ошибка", f"Ошибка при печати текстовой этикетки:\n{e}")

    def print_custom_text(self, text, font_size=24):
        """Печать этикетки с произвольным текстом (поддерживает несколько строк)"""
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import mm
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import tempfile
            import os
            import sys

            # Создаем временный файл
            temp_dir = tempfile.gettempdir()
            pdf_path = os.path.join(temp_dir, "text_label.pdf")

            # Создаем PDF (размер этикетки 58x40 мм)
            c = canvas.Canvas(pdf_path, pagesize=(58*mm, 40*mm))

            # Регистрируем шрифт с поддержкой кириллицы
            font_path = None
            possible_fonts = [
                "arial.ttf", "Arial.ttf", "ARIAL.TTF",
                "LiberationSans-Regular.ttf", "DejaVuSans.ttf",
                "calibri.ttf", "Calibri.ttf", "CALIBRI.TTF",
                "times.ttf", "Times.ttf", "TIMES.TTF"
            ]

            for font_name in possible_fonts:
                if os.path.exists(font_name):
                    font_path = font_name
                    break
                # Также проверим в системных каталогах
                system_font_paths = [
                    r"C:\Windows\Fonts\%s" % font_name,
                    r"/usr/share/fonts/truetype/dejavu/%s" % font_name,
                    r"/usr/share/fonts/truetype/liberation/%s" % font_name,
                    r"/System/Library/Fonts/%s" % font_name
                ]
                for path in system_font_paths:
                    if os.path.exists(path):
                        font_path = path
                        break
                if font_path:
                    break

            if font_path:
                try:
                    # Регистрируем шрифт для кириллицы
                    pdfmetrics.registerFont(TTFont('CyrillicFont', font_path))
                    font_name = 'CyrillicFont'
                except Exception as e:
                    logger.warning(f"Не удалось зарегистрировать шрифт {font_path}: {e}")
                    font_name = "Helvetica-Bold"
            else:
                font_name = "Helvetica-Bold"

            c.setFont(font_name, font_size)
            
            # Разбиваем текст на строки
            lines = text.split('\n')
            
            # Вычисляем высоту строки
            line_height = font_size * 1.2
            
            # Вычисляем общую высоту текста
            total_height = len(lines) * line_height
            
            # Начальная позиция Y (центрируем по вертикали)
            y_start = (40*mm - total_height) / 2 + (len(lines) - 1) * line_height
            
            # Рисуем каждую строку
            for line in lines:
                text_width = c.stringWidth(line, font_name, font_size)
                x_center = (58*mm - text_width) / 2
                c.drawString(x_center, y_start, line)
                y_start -= line_height
            
            c.save()

            # Печатаем PDF с использованием функции print_pdf
            result = print_pdf(pdf_path, use_acrobat=False)

            # Удаляем временный файл после печати
            try:
                os.remove(pdf_path)
            except:
                pass  # Игнорируем ошибки при удалении временного файла

            if not result:
                QMessageBox.warning(
                    self, "Предупреждение",
                    "Не удалось автоматически отправить этикетку на печать.\n"
                    "Файл этикетки был сохранён во временную папку.\n"
                    "Распечатайте его вручную из программы просмотра PDF."
                )

        except ImportError:
            logger.error("Модуль reportlab не установлен. Установите командой: pip install reportlab")
            QMessageBox.critical(self, "Ошибка", "Модуль reportlab не установлен.\nУстановите командой:\npip install reportlab")
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Ошибка при создании текстовой этикетки: {e}", exc_info=True)
            QMessageBox.critical(self, "Ошибка", f"Ошибка при создании текстовой этикетки:\n{e}")

    def open_label_settings(self):
        """Открытие диалога настроек печати этикеток"""
        try:
            from label_settings_dialog import LabelSettingsDialog
            dialog = LabelSettingsDialog(self)
            dialog.exec()
        except ImportError:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "Ошибка", "Модуль настроек этикеток не найден. Убедитесь, что файл label_settings_dialog.py находится в проекте.")
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Ошибка при открытии диалога настроек этикеток: {e}", exc_info=True)
            QMessageBox.critical(self, "Ошибка", f"Ошибка при открытии диалога настроек этикеток:\n{e}")
    
    def load_sku_data_from_db(self):
        """Загружает данные из общей базы данных в память"""
        sku_data = {}
        try:
            # Используем общую базу данных PostgreSQL
            from database import execute_query
            rows = execute_query("SELECT barcode, article, name FROM sku", fetchall=True)
            
            # Загружаем данные из базы
            for row in rows:
                barcode, article, name = row
                sku_data[barcode] = {'article': article, 'name': name}
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить данные из базы: {str(e)}")
        
        return sku_data
    
    def update_sku_table(self):
        """Актуализировать таблицу SKU из файла SKU.xlsx (асинхронно)"""
        try:
            import os
            sku_path = "SKU/SKU.xlsx"
            
            if not os.path.exists(sku_path):
                QMessageBox.critical(self, "Ошибка", f"Файл {sku_path} не найден!")
                return
            
            # Запускаем асинхронную обработку
            from PyQt6.QtCore import QThread, QObject, pyqtSignal
            
            class SKUWorker(QObject):
                finished = pyqtSignal(dict)
                error = pyqtSignal(str)
                
                def __init__(self, sku_path):
                    super().__init__()
                    self.sku_path = sku_path
                
                def run(self):
                    try:
                        from openpyxl import load_workbook
                        from database import execute_query
                        
                        # Удаляем все существующие данные из таблицы sku
                        execute_query("DELETE FROM sku")
                        
                        # Загружаем данные из Excel файла
                        wb = load_workbook(self.sku_path)
                        
                        try:
                            ws = wb["Лист1"]
                        except KeyError:
                            ws = wb.active
                        
                        # Получаем список всех PDF файлов
                        pdf_barcodes = set()
                        if os.path.exists("PDF"):
                            pdf_barcodes = {f[:-4] for f in os.listdir("PDF") if f.endswith('.pdf')}
                        
                        table_barcodes = set()
                        barcode_rows = {}
                        
                        for row in range(2, ws.max_row + 1):
                            barcode_cell = ws.cell(row=row, column=1)
                            if barcode_cell.value is None:
                                continue
                            barcode = str(barcode_cell.value)
                            table_barcodes.add(barcode)
                            barcode_rows[barcode] = row
                        
                        # Обновляем столбец "Этикетка (есть/нет)"
                        for barcode, row in barcode_rows.items():
                            label_status = "Есть" if barcode in pdf_barcodes else "Нет"
                            ws.cell(row=row, column=4, value=label_status)
                        
                        # Лист "Добавить" для штрихкодов из PDF, но не в SKU
                        try:
                            add_ws = wb["Добавить"]
                        except KeyError:
                            add_ws = wb.create_sheet("Добавить")
                            add_ws.cell(row=1, column=1, value="Штрихкод")
                            add_ws.cell(row=1, column=2, value="Артикул")
                            add_ws.cell(row=1, column=3, value="Наименование")
                            add_ws.cell(row=1, column=4, value="Этикетка (есть/нет)")
                        
                        next_add_row = 2
                        while next_add_row <= add_ws.max_row and add_ws.cell(row=next_add_row, column=1).value is not None:
                            next_add_row += 1
                        
                        missing_barcodes = pdf_barcodes - table_barcodes
                        for barcode in missing_barcodes:
                            add_ws.cell(row=next_add_row, column=1, value=barcode)
                            add_ws.cell(row=next_add_row, column=4, value="Есть")
                            next_add_row += 1
                        
                        # Сохраняем Excel
                        wb.save(self.sku_path)
                        
                        # Вставляем данные в БД
                        records_with_labels = 0
                        for row in range(2, ws.max_row + 1):
                            barcode_cell = ws.cell(row=row, column=1)
                            if barcode_cell.value is None:
                                continue
                            
                            barcode = str(barcode_cell.value)
                            article = ws.cell(row=row, column=2).value or ""
                            name = ws.cell(row=row, column=3).value or ""
                            
                            execute_query(
                                "INSERT INTO sku (barcode, article, name) VALUES (%s, %s, %s) ON CONFLICT (barcode) DO UPDATE SET article = EXCLUDED.article, name = EXCLUDED.name",
                                (barcode, article, name)
                            )
                            
                            if barcode in pdf_barcodes:
                                records_with_labels += 1
                        
                        self.finished.emit({
                            'records_with_labels': records_with_labels,
                            'missing_barcodes_count': len(missing_barcodes),
                        })
                    except Exception as e:
                        self.error.emit(str(e))
            
            # Создаём и запускаем worker в отдельном потоке
            self._sku_thread = QThread()
            self._sku_worker = SKUWorker(sku_path)
            self._sku_worker.moveToThread(self._sku_thread)
            self._sku_thread.started.connect(self._sku_worker.run)
            self._sku_worker.finished.connect(self._on_sku_update_finished)
            self._sku_worker.error.connect(self._on_sku_update_error)
            self._sku_worker.finished.connect(self._sku_thread.quit)
            self._sku_worker.error.connect(self._sku_thread.quit)
            self._sku_thread.finished.connect(self._sku_thread.deleteLater)
            self._sku_worker.finished.connect(self._sku_worker.deleteLater)
            self._sku_thread.start()
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось запустить обновление таблицы: {str(e)}")
    
    def _on_sku_update_finished(self, result):
        """Обработка успешного обновления SKU"""
        QMessageBox.information(
            self, "Успех",
            f"Таблица SKU успешно обновлена из файла Excel!\n"
            f"Количество штрихкодов с этикетками: {result['records_with_labels']}\n"
            f"Количество новых штрихкодов для добавления: {result['missing_barcodes_count']}"
        )
    
    def _on_sku_update_error(self, error_msg):
        """Обработка ошибки обновления SKU"""
        QMessageBox.critical(self, "Ошибка", f"Не удалось обновить таблицу: {error_msg}")
    
    def print_product_label(self):
        """Печать этикетки на товар"""
        # Цикл для многократного ввода штрихкода и печати этикеток
        while True:
            # Создаем диалог для ввода штрихкода и количества
            dialog = BarcodeQuantityDialog(self)

            # Если пользователь нажал Cancel или закрыл окно, выходим из цикла
            if dialog.exec() != QDialog.DialogCode.Accepted:
                break

            barcode = dialog.get_barcode()
            quantity = dialog.get_quantity()

            # Проверяем, введён ли штрихкод
            if not barcode:
                QMessageBox.warning(self, "Ошибка", "Введите штрихкод товара!")
                continue

            # Проверяем, существует ли PDF файл с таким штрихкодом
            pdf_path = f"PDF/{barcode}.pdf"

            if os.path.exists(pdf_path):
                # Печатаем PDF файл нужное количество раз
                success_count = 0
                for i in range(quantity):
                    if print_pdf(pdf_path, use_acrobat=False):
                        success_count += 1

                if success_count == 0:
                    QMessageBox.critical(self, "Ошибка", f"Не удалось распечатать PDF файл для штрихкода {barcode}!")
                # После успешной печати продолжаем цикл, чтобы снова запросить штрихкод
            else:
                QMessageBox.warning(self, "Ошибка", f"PDF файл для штрихкода {barcode} не найден!")
        
    def print_box_label(self):
        """Печать этикетки на коробку (независимо от поставок и коробок)"""
        # Цикл для многократного ввода штрихкода и печати этикеток
        while True:
            # Запрос штрихкода через диалоговое окно
            barcode, ok = QInputDialog.getText(self, "Штрихкод", "Введите штрихкод:")

            # Если пользователь нажал Cancel или закрыл окно, выходим из цикла
            if not ok or not barcode:
                break

            # Загружаем данные SKU из базы данных
            sku_data = self.load_sku_data_from_db()

            # Проверяем, пуста ли таблица SKU
            if not sku_data:
                reply = QMessageBox.question(
                    self, "Таблица SKU пуста",
                    "Таблица SKU пуста или не загружена.\n\n"
                    "Хотите актуализировать таблицу SKU из файла Excel?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                if reply == QMessageBox.StandardButton.Yes:
                    self.update_sku_table()
                    sku_data = self.load_sku_data_from_db()
                else:
                    break

            # Ищем артикул в загруженных данных
            if barcode in sku_data:
                data = sku_data[barcode]
                article = data['article']
                name = data['name']
                # Создаем и печатаем этикетку с использованием данных из памяти
                printer = QPrinter()
                custom_size = QPageSize(QSizeF(58, 40), QPageSize.Unit.Millimeter)
                printer.setPageSize(custom_size)
                printer.setPageOrientation(QPageLayout.Orientation.Portrait)

                if create_and_print_box_label(printer, str(barcode), str(article), str(name)):
                    # После успешной печати продолжаем цикл, чтобы снова запросить штрихкод
                    continue
                else:
                    QMessageBox.critical(self, "Ошибка", f"Не удалось создать и напечатать этикетку для штрихкода {barcode}!")
            else:
                QMessageBox.warning(self, "Ошибка", f"Данные для штрихкода {barcode} не найдены в таблице SKU!")

            # Спрашиваем пользователя, хочет ли он попробовать снова
            reply = QMessageBox.question(self, "Повторить?", "Хотите ввести другой штрихкод?",
                                       QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply != QMessageBox.StandardButton.Yes:
                break